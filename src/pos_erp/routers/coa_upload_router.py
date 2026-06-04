"""COA Upload Router — Onboarding bulk import with validation & auto-mapping."""

from __future__ import annotations
import re
import io
from typing import Optional
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from pos_erp.db import fetch_all, fetch_one, execute, execute_returning

router = APIRouter(prefix="/api/coa/upload", tags=["COA Upload"])

VALID_TYPES = {"ASSET", "LIABILITY", "EQUITY", "REVENUE", "EXPENSE"}
CODE_PATTERN = re.compile(r"^\d+(\.\d+)*$")


# ── Models ───────────────────────────────────────────────────────────────────

class COARow(BaseModel):
    level: int
    parent_code: Optional[str] = None
    account_code: str
    account_name: str
    account_type: str
    is_active: bool = True


class ValidateRequest(BaseModel):
    rows: list[COARow]


class ApplyRequest(BaseModel):
    rows: list[COARow]
    skip_invalid: bool = True


class MappingConfirmRequest(BaseModel):
    confirmed: bool = True


# ── Template Download ────────────────────────────────────────────────────────

@router.get("/template/download")
def download_template():
    """Generate Excel template with 5 sheets."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="C9A96E", end_color="C9A96E", fill_type="solid")
    example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    headers = ["level", "parent_code", "account_code", "account_name", "account_type", "is_active"]

    def setup_header(ws, title=None):
        if title:
            ws.title = title
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 35
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 10

    def add_example_rows(ws, rows, start_row=2):
        for i, row_data in enumerate(rows):
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=start_row + i, column=col, value=val)
                cell.fill = example_fill

    # Sheet 1: Template kosong
    ws1 = wb.active
    setup_header(ws1, "Template")

    # Sheet 2: Petunjuk
    ws2 = wb.create_sheet("Petunjuk")
    ws2.column_dimensions["A"].width = 80
    instructions = [
        "PETUNJUK PENGISIAN TEMPLATE COA (Chart of Accounts)",
        "",
        "1. KOLOM 'level':",
        "   Isi dengan angka 1, 2, 3, atau 4.",
        "   - Level 1 = Kelompok Akun (contoh: Aset, Kewajiban, Pendapatan)",
        "   - Level 2 = Golongan (contoh: Aset Lancar, Pendapatan Jasa)",
        "   - Level 3 = Sub-golongan (contoh: Kas & Bank, Pendapatan Layanan Eyelash)",
        "   - Level 4 = Akun Detail (contoh: Kas Tunai, Bank BCA) — ini yang dipakai transaksi",
        "",
        "2. KOLOM 'parent_code':",
        "   - Level 1: kosongkan (tidak punya parent)",
        "   - Level 2: isi kode akun Level 1 (contoh: 1)",
        "   - Level 3: isi kode akun Level 2 (contoh: 1.1)",
        "   - Level 4: isi kode akun Level 3 (contoh: 1.1.1)",
        "",
        "3. KOLOM 'account_code':",
        "   Format: 1 atau 1.1 atau 1.1.1 atau 1.1.1.1 (pakai titik, bukan strip)",
        "   Setiap kode harus unik, tidak boleh ada yang sama.",
        "",
        "4. KOLOM 'account_name':",
        "   Nama akun yang jelas dan konsisten.",
        "   Untuk Level 4 (akun detail), nama harus sesuai dengan nama item di POS.",
        "",
        "5. KOLOM 'account_type':",
        "   Pilih salah satu: ASSET, LIABILITY, EQUITY, REVENUE, EXPENSE",
        "   Level 4 harus konsisten dengan parent-nya (misal: di bawah Pendapatan = REVENUE)",
        "",
        "6. KOLOM 'is_active':",
        "   TRUE = akun aktif, FALSE = akun nonaktif",
        "",
        "TIPS:",
        "- Lihat Sheet 3-5 untuk contoh struktur yang benar",
        "- Sistem akan otomatis mapping akun Level 4 ke transaksi POS",
        "- Setelah upload pertama kali, akun baru ditambahkan satu per satu (bukan upload ulang)",
    ]
    for i, line in enumerate(instructions, 1):
        cell = ws2.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=14)

    # Sheet 3: Contoh Aset
    ws3 = wb.create_sheet("Contoh - Aset")
    setup_header(ws3)
    add_example_rows(ws3, [
        [1, None, "1", "Aset", "ASSET", True],
        [2, "1", "1.1", "Aset Lancar", "ASSET", True],
        [3, "1.1", "1.1.1", "Kas & Bank", "ASSET", True],
        [4, "1.1.1", "1.1.1.1", "Kas Tunai", "ASSET", True],
        [4, "1.1.1", "1.1.1.2", "QRIS Clearing", "ASSET", True],
        [4, "1.1.1", "1.1.1.3", "Bank BCA", "ASSET", True],
        [3, "1.1", "1.1.2", "Piutang Usaha", "ASSET", True],
        [4, "1.1.2", "1.1.2.1", "Piutang Customer", "ASSET", True],
    ])

    # Sheet 4: Contoh Pendapatan
    ws4 = wb.create_sheet("Contoh - Pendapatan")
    setup_header(ws4)
    add_example_rows(ws4, [
        [1, None, "4", "Pendapatan", "REVENUE", True],
        [2, "4", "4.1", "Pendapatan Jasa", "REVENUE", True],
        [3, "4.1", "4.1.1", "Pendapatan Layanan Eyelash", "REVENUE", True],
        [4, "4.1.1", "4.1.1.1", "Eyelash Extension Classic", "REVENUE", True],
        [4, "4.1.1", "4.1.1.2", "Eyelash Extension Volume", "REVENUE", True],
        [4, "4.1.1", "4.1.1.3", "Eyelash Removal", "REVENUE", True],
    ])

    # Sheet 5: Contoh Diskon & Rounding
    ws5 = wb.create_sheet("Contoh - Diskon & Rounding")
    setup_header(ws5)
    add_example_rows(ws5, [
        [1, None, "5", "Beban", "EXPENSE", True],
        [2, "5", "5.1", "Beban Operasional", "EXPENSE", True],
        [3, "5.1", "5.1.1", "Diskon & Potongan", "EXPENSE", True],
        [4, "5.1.1", "5.1.1.1", "Diskon Member", "EXPENSE", True],
        [4, "5.1.1", "5.1.1.2", "Selisih Pembulatan", "EXPENSE", True],
    ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_coa_beauty_shine.xlsx"},
    )


# ── Validate ─────────────────────────────────────────────────────────────────

@router.post("/validate")
def validate_upload(data: ValidateRequest):
    """Validate COA rows before apply. Returns valid/invalid/warnings."""
    rows = data.rows
    existing = fetch_all("SELECT account_code, account_name FROM chart_of_account")
    existing_codes = {r["account_code"] for r in existing}
    existing_names = {r["account_name"] for r in existing}

    valid = []
    invalid = []
    warnings = []

    seen_codes = set()
    code_to_row = {}
    for i, row in enumerate(rows):
        code_to_row[row.account_code] = row

    for i, row in enumerate(rows):
        errors = []

        # Format validation
        if not CODE_PATTERN.match(row.account_code):
            errors.append(f"Format kode '{row.account_code}' tidak valid (harus pakai titik, contoh: 1.1.1.1)")

        if not row.account_name or not row.account_name.strip():
            errors.append("Nama akun tidak boleh kosong")

        if row.account_type not in VALID_TYPES:
            errors.append(f"Tipe '{row.account_type}' tidak valid (harus: ASSET/LIABILITY/EQUITY/REVENUE/EXPENSE)")

        if row.level not in (1, 2, 3, 4):
            errors.append(f"Level harus 1-4, bukan {row.level}")

        # Uniqueness in upload
        if row.account_code in seen_codes:
            errors.append(f"Kode '{row.account_code}' duplikat dalam file upload")
        seen_codes.add(row.account_code)

        # Duplicate with existing
        if row.account_code in existing_codes:
            errors.append(f"Kode '{row.account_code}' sudah ada di sistem")

        # Hierarchy validation
        if row.level == 1 and row.parent_code:
            errors.append("Level 1 tidak boleh punya parent_code")
        elif row.level >= 2:
            if not row.parent_code:
                errors.append(f"Level {row.level} harus punya parent_code")
            elif row.parent_code not in code_to_row and row.parent_code not in existing_codes:
                errors.append(f"Parent '{row.parent_code}' tidak ditemukan")

        # Level 4 type consistency
        if row.level == 4 and row.parent_code and row.parent_code in code_to_row:
            parent = code_to_row[row.parent_code]
            if parent.account_type != row.account_type:
                errors.append(f"Tipe '{row.account_type}' tidak konsisten dengan parent '{parent.account_name}' ({parent.account_type})")

        # Warning: duplicate name different code
        if not errors and row.account_name in existing_names:
            warnings.append({
                "row": i + 1,
                "account_code": row.account_code,
                "account_name": row.account_name,
                "warning": f"Nama '{row.account_name}' sudah ada di sistem dengan kode berbeda",
            })

        row_dict = row.model_dump()
        if errors:
            invalid.append({**row_dict, "row": i + 1, "errors": errors})
        else:
            valid.append({**row_dict, "row": i + 1})

    return {
        "valid": valid,
        "invalid": invalid,
        "warnings": warnings,
        "summary": {
            "total": len(rows),
            "valid_count": len(valid),
            "invalid_count": len(invalid),
            "warning_count": len(warnings),
        },
    }


# ── Apply ────────────────────────────────────────────────────────────────────

@router.post("/apply")
def apply_upload(data: ApplyRequest):
    """Apply validated COA rows to database."""
    existing = fetch_all("SELECT account_code FROM chart_of_account")
    existing_codes = {r["account_code"] for r in existing}

    inserted = 0
    skipped = 0
    errors_list = []

    for row in data.rows:
        if row.account_code in existing_codes:
            skipped += 1
            continue

        try:
            execute(
                """INSERT INTO chart_of_account (account_code, account_name, account_type, parent_code, level, is_active)
                   VALUES (%s, %s, %s, %s, %s, %s)""",
                (row.account_code, row.account_name, row.account_type,
                 row.parent_code, row.level, row.is_active),
            )
            inserted += 1
        except Exception as e:
            errors_list.append({"account_code": row.account_code, "error": str(e)})

    # Run auto-mapping after insert
    mapping_result = _auto_map_accounts()

    return {
        "inserted": inserted,
        "skipped": skipped,
        "errors": errors_list,
        "mapping": mapping_result,
    }


# ── Mapping Preview ──────────────────────────────────────────────────────────

@router.get("/mapping-preview")
def mapping_preview():
    """Preview auto-mapping results for Level 4 accounts."""
    return _auto_map_accounts()


@router.post("/mapping-confirm")
def confirm_mapping(data: MappingConfirmRequest):
    """Confirm and apply auto-mapping to account_mapping table."""
    if not data.confirmed:
        return {"confirmed": False}

    result = _auto_map_accounts()
    applied = 0

    for m in result.get("autoMapped", []):
        try:
            existing = fetch_one(
                "SELECT id FROM account_mapping WHERE module = %s AND transaction_type = %s",
                (m["module"], m["transaction_type"]),
            )
            if existing:
                execute(
                    "UPDATE account_mapping SET debit_account = %s, credit_account = %s, description = %s WHERE id = %s",
                    (m.get("debit_account", ""), m.get("credit_account", ""), m.get("description", ""), existing["id"]),
                )
            else:
                execute(
                    """INSERT INTO account_mapping (module, transaction_type, debit_account, credit_account, description)
                       VALUES (%s, %s, %s, %s, %s)""",
                    (m["module"], m["transaction_type"], m.get("debit_account", ""), m.get("credit_account", ""), m.get("description", "")),
                )
            applied += 1
        except Exception:
            pass

    return {"confirmed": True, "applied": applied, "total": len(result.get("autoMapped", []))}


# ── Auto-mapping logic ───────────────────────────────────────────────────────

def _auto_map_accounts():
    """Detect mapping candidates from Level 4 accounts."""
    accounts = fetch_all(
        "SELECT account_code, account_name, account_type FROM chart_of_account WHERE level = 4 AND is_active = true"
    )

    auto_mapped = []
    needs_review = []

    for acc in accounts:
        name_lower = acc["account_name"].lower()
        code = acc["account_code"]
        atype = acc["account_type"]

        mapped = False

        # Cash mapping
        if atype == "ASSET" and any(k in name_lower for k in ["kas", "cash", "tunai"]):
            auto_mapped.append({
                "account_code": code, "account_name": acc["account_name"],
                "module": "pos", "transaction_type": "cash_payment",
                "debit_account": code, "credit_account": "",
                "description": f"Kas — {acc['account_name']}",
                "reason": "Detected: kas/cash/tunai in ASSET account",
            })
            mapped = True

        # QRIS mapping
        if atype == "ASSET" and any(k in name_lower for k in ["qris", "clearing", "e-wallet", "ewallet"]):
            auto_mapped.append({
                "account_code": code, "account_name": acc["account_name"],
                "module": "pos", "transaction_type": "qris_payment",
                "debit_account": code, "credit_account": "",
                "description": f"QRIS — {acc['account_name']}",
                "reason": "Detected: QRIS/clearing in ASSET account",
            })
            mapped = True

        # Bank/Transfer mapping
        if atype == "ASSET" and any(k in name_lower for k in ["bank", "transfer", "bca", "mandiri", "bni", "bri"]):
            auto_mapped.append({
                "account_code": code, "account_name": acc["account_name"],
                "module": "pos", "transaction_type": "bank_payment",
                "debit_account": code, "credit_account": "",
                "description": f"Bank — {acc['account_name']}",
                "reason": "Detected: bank/transfer in ASSET account",
            })
            mapped = True

        # Revenue → service mapping
        if atype == "REVENUE":
            auto_mapped.append({
                "account_code": code, "account_name": acc["account_name"],
                "module": "pos", "transaction_type": f"service_{code.replace('.', '_')}",
                "debit_account": "", "credit_account": code,
                "description": f"Revenue — {acc['account_name']}",
                "reason": "Detected: REVENUE Level 4 → service mapping",
            })
            mapped = True

        # Discount mapping
        if atype == "EXPENSE" and any(k in name_lower for k in ["diskon", "potongan", "discount"]):
            auto_mapped.append({
                "account_code": code, "account_name": acc["account_name"],
                "module": "pos", "transaction_type": "discount",
                "debit_account": code, "credit_account": "",
                "description": f"Diskon — {acc['account_name']}",
                "reason": "Detected: diskon/potongan in EXPENSE account",
            })
            mapped = True

        # Rounding mapping
        if atype == "EXPENSE" and any(k in name_lower for k in ["pembulatan", "rounding", "selisih"]):
            auto_mapped.append({
                "account_code": code, "account_name": acc["account_name"],
                "module": "pos", "transaction_type": "rounding",
                "debit_account": code, "credit_account": "",
                "description": f"Rounding — {acc['account_name']}",
                "reason": "Detected: pembulatan/rounding in EXPENSE account",
            })
            mapped = True

        if not mapped:
            needs_review.append({
                "account_code": code,
                "account_name": acc["account_name"],
                "account_type": atype,
                "reason": "Tidak terdeteksi otomatis — perlu mapping manual",
            })

    return {
        "autoMapped": auto_mapped,
        "needsReview": needs_review,
        "failed": [],
    }
