"""Reporting & Dashboard Router - Beauty & Shine ERP.

Handles:
- Owner Dashboard (KPI, exceptions, approvals)
- Sales reports
- Inventory reports
- Financial reports
- Treatment reports
- Staff performance
- CSV export
"""
from __future__ import annotations

import csv
import io
from datetime import date, datetime, timedelta
from typing import Optional

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from pos_erp.db import fetch_all, fetch_one, execute

router = APIRouter(prefix="/api/reporting", tags=["Reporting & Dashboard"])


# ── Owner Dashboard ───────────────────────────────────────────────

@router.get("/dashboard")
def get_owner_dashboard(branch_id: str = ""):
    """Get owner dashboard with KPIs."""
    today = date.today()
    month_start = today.replace(day=1)
    
    # Sales KPIs
    sales_conditions = ["t.status = 'paid'"]
    sales_params = []
    if branch_id:
        sales_conditions.append("t.branch_id = %s")
        sales_params.append(branch_id)
    sales_where = " AND ".join(sales_conditions)
    
    # Today's sales
    today_sales = fetch_one(
        f"SELECT COALESCE(SUM(total), 0) as amount, count(*) as count FROM pos_transaction t WHERE {sales_where} AND DATE(t.created_at) = %s",
        (*sales_params, today)
    )
    
    # Month sales
    month_sales = fetch_one(
        f"SELECT COALESCE(SUM(total), 0) as amount, count(*) as count FROM pos_transaction t WHERE {sales_where} AND DATE(t.created_at) >= %s",
        (*sales_params, month_start)
    )
    
    # Pending sync
    pending_sync = fetch_one("SELECT count(*) as count FROM sync_queue WHERE status IN ('pending', 'retry')")
    
    # Open exceptions
    open_exceptions = fetch_one("SELECT count(*) as count FROM exception_queue WHERE status = 'open'")
    
    # Pending approvals
    pending_approvals = fetch_one("SELECT count(*) as count FROM approval_matrix WHERE status = 'pending'")
    
    # Low stock count
    low_stock = fetch_one("SELECT count(*) as count FROM stock_card WHERE balance < 10")
    
    # Top treatments
    top_treatments = fetch_all(
        """SELECT ti.item_name, count(*) as count, SUM(ti.total) as revenue
           FROM pos_transaction_item ti
           JOIN pos_transaction t ON ti.transaction_id = t.id
           WHERE ti.item_type = 'treatment' AND t.status = 'paid'
           GROUP BY ti.item_name ORDER BY count DESC LIMIT 5"""
    )
    
    # Recent transactions
    recent_txns = fetch_all(
        """SELECT t.doc_key, t.customer_name, t.total, t.status, t.created_at
           FROM pos_transaction t
           ORDER BY t.created_at DESC LIMIT 10"""
    )
    
    return {
        "today": {
            "sales_amount": float(today_sales["amount"] or 0),
            "transaction_count": today_sales["count"],
        },
        "month": {
            "sales_amount": float(month_sales["amount"] or 0),
            "transaction_count": month_sales["count"],
        },
        "alerts": {
            "pending_sync": pending_sync["count"] if pending_sync else 0,
            "open_exceptions": open_exceptions["count"] if open_exceptions else 0,
            "pending_approvals": pending_approvals["count"] if pending_approvals else 0,
            "low_stock": low_stock["count"] if low_stock else 0,
        },
        "top_treatments": top_treatments,
        "recent_transactions": recent_txns,
        "sales_trend": fetch_all(
            f"""SELECT DATE(t.created_at) as date,
                COUNT(*) as transactions,
                COALESCE(SUM(t.total), 0) as revenue
            FROM pos_transaction t
            WHERE {sales_where} AND DATE(t.created_at) >= %s
            GROUP BY DATE(t.created_at)
            ORDER BY date""",
            (*sales_params, today - timedelta(days=13))
        ),
        "sales_by_payment": fetch_all(
            """SELECT pm.name as method, COUNT(*) as count, COALESCE(SUM(t.total), 0) as amount
            FROM pos_transaction t
            LEFT JOIN payment_method pm ON t.payment_method_id = pm.id
            WHERE t.status = 'paid' AND DATE(t.created_at) >= %s
            GROUP BY pm.name ORDER BY amount DESC""",
            (month_start,)
        ),
        "low_stock_items": fetch_all(
            """SELECT p.name, p.sku, sc.balance, sc.last_movement_date
            FROM stock_card sc JOIN product p ON sc.product_id = p.id
            WHERE sc.balance < 10 AND p.is_active = true
            ORDER BY sc.balance LIMIT 5"""
        ),
    }


# ── Sales Reports ─────────────────────────────────────────────────

@router.get("/sales/daily")
def get_daily_sales_report(branch_id: str = "", date_from: str = "", date_to: str = ""):
    """Get daily sales report."""
    conditions = ["t.status = 'paid'"]
    params = []
    if branch_id:
        conditions.append("t.branch_id = %s")
        params.append(branch_id)
    if date_from:
        conditions.append("DATE(t.created_at) >= %s")
        params.append(date_from)
    if date_to:
        conditions.append("DATE(t.created_at) <= %s")
        params.append(date_to)
    where = " AND ".join(conditions)
    
    rows = fetch_all(
        f"""SELECT DATE(t.created_at) as sale_date,
                   count(*) as transaction_count,
                   SUM(t.subtotal) as subtotal,
                   SUM(t.discount) as discount,
                   SUM(t.tax) as tax,
                   SUM(t.total) as total
            FROM pos_transaction t
            WHERE {where}
            GROUP BY DATE(t.created_at)
            ORDER BY sale_date DESC""",
        tuple(params)
    )
    return {"items": rows}

@router.get("/sales/by-treatment")
def get_sales_by_treatment(branch_id: str = "", date_from: str = "", date_to: str = ""):
    """Get sales breakdown by treatment."""
    conditions = ["t.status = 'paid'", "ti.item_type = 'treatment'"]
    params = []
    if branch_id:
        conditions.append("t.branch_id = %s")
        params.append(branch_id)
    if date_from:
        conditions.append("DATE(t.created_at) >= %s")
        params.append(date_from)
    if date_to:
        conditions.append("DATE(t.created_at) <= %s")
        params.append(date_to)
    where = " AND ".join(conditions)
    
    rows = fetch_all(
        f"""SELECT ti.item_name, count(*) as count, SUM(ti.total) as revenue
            FROM pos_transaction_item ti
            JOIN pos_transaction t ON ti.transaction_id = t.id
            WHERE {where}
            GROUP BY ti.item_name
            ORDER BY revenue DESC""",
        tuple(params)
    )
    return {"items": rows}

@router.get("/sales/by-payment")
def get_sales_by_payment_method(branch_id: str = "", date_from: str = "", date_to: str = ""):
    """Get sales breakdown by payment method."""
    conditions = ["t.status = 'paid'"]
    params = []
    if branch_id:
        conditions.append("t.branch_id = %s")
        params.append(branch_id)
    if date_from:
        conditions.append("DATE(t.created_at) >= %s")
        params.append(date_from)
    if date_to:
        conditions.append("DATE(t.created_at) <= %s")
        params.append(date_to)
    where = " AND ".join(conditions)
    
    rows = fetch_all(
        f"""SELECT pm.name as payment_method, count(*) as count, SUM(t.total) as total
            FROM pos_transaction t
            LEFT JOIN payment_method pm ON t.payment_method_id = pm.id
            WHERE {where}
            GROUP BY pm.name
            ORDER BY total DESC""",
        tuple(params)
    )
    return {"items": rows}


# ── Treatment Reports ─────────────────────────────────────────────

@router.get("/treatments/summary")
def get_treatment_summary(branch_id: str = "", date_from: str = "", date_to: str = ""):
    """Get treatment summary report."""
    conditions = []
    params = []
    if branch_id:
        conditions.append("tr.transaction_id IN (SELECT id FROM pos_transaction WHERE branch_id = %s)")
        params.append(branch_id)
    if date_from:
        conditions.append("tr.start_time >= %s")
        params.append(date_from)
    if date_to:
        conditions.append("tr.start_time <= %s")
        params.append(date_to + " 23:59:59")
    where = "WHERE " + " AND ".join(conditions) if conditions else ""
    
    rows = fetch_all(
        f"""SELECT t.name as treatment_name, tc.name as category,
                   count(*) as total_sessions,
                   SUM(CASE WHEN tr.status = 'completed' THEN 1 ELSE 0 END) as completed,
                   SUM(CASE WHEN tr.status = 'cancelled' THEN 1 ELSE 0 END) as cancelled
            FROM treatment_record tr
            JOIN treatment t ON tr.treatment_id = t.id
            LEFT JOIN treatment_category tc ON t.category_id = tc.id
            {where}
            GROUP BY t.name, tc.name
            ORDER BY total_sessions DESC""",
        tuple(params)
    )
    return {"items": rows}

@router.get("/treatments/therapist-performance")
def get_therapist_performance(branch_id: str = "", date_from: str = "", date_to: str = ""):
    """Get therapist performance report."""
    conditions = ["tr.status = 'completed'"]
    params = []
    if branch_id:
        conditions.append("tr.transaction_id IN (SELECT id FROM pos_transaction WHERE branch_id = %s)")
        params.append(branch_id)
    if date_from:
        conditions.append("tr.start_time >= %s")
        params.append(date_from)
    if date_to:
        conditions.append("tr.start_time <= %s")
        params.append(date_to + " 23:59:59")
    where = " AND ".join(conditions)
    
    rows = fetch_all(
        f"""SELECT u.full_name as therapist_name,
                   count(*) as total_sessions,
                   SUM(ti.total) as revenue_generated
            FROM treatment_record tr
            JOIN app_user u ON tr.therapist_id = u.id
            LEFT JOIN pos_transaction_item ti ON tr.transaction_id = ti.transaction_id AND tr.treatment_id = ti.item_id
            WHERE {where}
            GROUP BY u.full_name
            ORDER BY revenue_generated DESC""",
        tuple(params)
    )
    return {"items": rows}


# ── Inventory Reports ─────────────────────────────────────────────

@router.get("/inventory/stock-summary")
def get_stock_summary(branch_id: str = ""):
    """Get inventory stock summary."""
    extra = "WHERE sc.branch_id = %s" if branch_id else ""
    params = (branch_id,) if branch_id else ()
    
    rows = fetch_all(
        f"""SELECT p.name as product_name, p.sku, p.unit,
                   sc.qty_in, sc.qty_out, sc.balance,
                   sc.last_movement_date
            FROM stock_card sc
            JOIN product p ON sc.product_id = p.id
            {extra}
            ORDER BY sc.balance ASC""",
        params
    )
    
    total_items = len(rows)
    total_value = sum(float(r["balance"]) for r in rows)
    low_stock_count = sum(1 for r in rows if float(r["balance"]) < 10)
    
    return {
        "items": rows,
        "summary": {
            "total_items": total_items,
            "total_quantity": total_value,
            "low_stock_count": low_stock_count,
        }
    }

@router.get("/inventory/movement-summary")
def get_movement_summary(branch_id: str = "", date_from: str = "", date_to: str = ""):
    """Get stock movement summary."""
    conditions = []
    params = []
    if branch_id:
        conditions.append("sm.branch_id = %s")
        params.append(branch_id)
    if date_from:
        conditions.append("DATE(sm.created_at) >= %s")
        params.append(date_from)
    if date_to:
        conditions.append("DATE(sm.created_at) <= %s")
        params.append(date_to)
    where = "WHERE " + " AND ".join(conditions) if conditions else ""
    
    rows = fetch_all(
        f"""SELECT DATE(sm.created_at) as movement_date,
                   sm.movement_type,
                   count(*) as count,
                   SUM(sm.qty) as total_qty
            FROM stock_movement sm
            {where}
            GROUP BY DATE(sm.created_at), sm.movement_type
            ORDER BY movement_date DESC""",
        tuple(params)
    )
    return {"items": rows}

@router.get("/inventory/batch-expiry")
def get_batch_expiry_report(days: int = 30):
    """Get batches expiring within N days."""
    cutoff = date.today() + timedelta(days=days)
    rows = fetch_all(
        """SELECT pb.*, p.name as product_name, b.name as branch_name
           FROM product_batch pb
           JOIN product p ON pb.product_id = p.id
           LEFT JOIN branch b ON pb.branch_id = b.id
           WHERE pb.expiry_date <= %s AND pb.qty > 0
           ORDER BY pb.expiry_date ASC""",
        (cutoff,)
    )
    return {"items": rows, "cutoff_date": cutoff}


# ── Financial Reports ─────────────────────────────────────────────

@router.get("/finance/summary")
def get_finance_summary(branch_id: str = ""):
    """Get financial summary."""
    conditions = []
    params = []
    if branch_id:
        conditions.append("gl.branch_id = %s")
        params.append(branch_id)
    where = "WHERE " + " AND ".join(conditions) if conditions else ""
    
    # Revenue
    revenue = fetch_one(
        f"SELECT COALESCE(SUM(credit) - SUM(debit), 0) as amount FROM general_ledger gl JOIN chart_of_account coa ON gl.account_code = coa.account_code {where} {'AND' if where else 'WHERE'} coa.account_type = 'Revenue'",
        tuple(params)
    )
    
    # Expenses
    expenses = fetch_one(
        f"SELECT COALESCE(SUM(debit) - SUM(credit), 0) as amount FROM general_ledger gl JOIN chart_of_account coa ON gl.account_code = coa.account_code {where} {'AND' if where else 'WHERE'} coa.account_type = 'Expense'",
        tuple(params)
    )
    
    # AP outstanding
    ap_outstanding = fetch_one(
        "SELECT COALESCE(SUM(amount - paid_amount), 0) as amount FROM accounts_payable WHERE status IN ('open', 'partial')" + (" AND branch_id = %s" if branch_id else ""),
        (branch_id,) if branch_id else ()
    )
    
    # Bank balances
    bank_balances = fetch_all(
        "SELECT bank_name, account_no, balance FROM bank_account WHERE is_active = true" + (" AND branch_id = %s" if branch_id else ""),
        (branch_id,) if branch_id else ()
    )
    
    revenue_amount = float(revenue["amount"] or 0)
    expense_amount = float(expenses["amount"] or 0)
    
    return {
        "revenue": revenue_amount,
        "expenses": expense_amount,
        "net_profit": revenue_amount - expense_amount,
        "ap_outstanding": float(ap_outstanding["amount"] or 0),
        "bank_balances": bank_balances,
    }


# ── Exception Reports ─────────────────────────────────────────────

@router.get("/exceptions")
def get_exception_report(status: str = "open", limit: int = 50):
    """Get exception report."""
    rows = fetch_all(
        """SELECT * FROM exception_queue WHERE status = %s ORDER BY created_at DESC LIMIT %s""",
        (status, limit)
    )
    return {"items": rows}


# ── Audit Reports ─────────────────────────────────────────────────

@router.get("/audit")
def get_audit_report(module: str = "", action: str = "", limit: int = 100):
    """Get audit trail report."""
    conditions = []
    params = []
    if module:
        conditions.append("module = %s")
        params.append(module)
    if action:
        conditions.append("action = %s")
        params.append(action)
    where = "WHERE " + " AND ".join(conditions) if conditions else ""
    params.append(limit)
    
    rows = fetch_all(
        f"SELECT * FROM audit_trail {where} ORDER BY timestamp DESC LIMIT %s",
        tuple(params)
    )
    return {"items": rows}


# ── Staff Performance ─────────────────────────────────────────────

@router.get("/staff/therapist-performance")
def get_therapist_performance_v2(branch_id: str = "", date_from: str = "", date_to: str = ""):
    """Get detailed therapist performance report with revenue metrics."""
    conditions = ["tr.status = 'completed'"]
    params = []
    if branch_id:
        conditions.append("t.branch_id = %s")
        params.append(branch_id)
    if date_from:
        conditions.append("DATE(tr.start_time) >= %s")
        params.append(date_from)
    if date_to:
        conditions.append("DATE(tr.start_time) <= %s")
        params.append(date_to)
    where = " AND ".join(conditions)

    rows = fetch_all(
        f"""SELECT u.full_name as therapist_name,
                   COUNT(DISTINCT tr.id) as total_sessions,
                   COUNT(DISTINCT t.id) as total_transactions,
                   COALESCE(SUM(ti.total), 0) as total_revenue,
                   COALESCE(AVG(ti.total), 0) as avg_revenue_per_session
            FROM treatment_record tr
            JOIN app_user u ON tr.therapist_id = u.id
            LEFT JOIN pos_transaction t ON tr.transaction_id = t.id
            LEFT JOIN pos_transaction_item ti ON tr.transaction_id = ti.transaction_id
            WHERE {where}
            GROUP BY u.full_name
            ORDER BY total_revenue DESC""",
        tuple(params),
    )
    return {"items": rows}


# ── CSV Export ────────────────────────────────────────────────────

@router.get("/export/csv")
def export_report_csv(
    report_type: str = "sales",
    date_from: str = "",
    date_to: str = "",
    branch_id: str = "",
):
    """Export report data as CSV file."""
    conditions = ["t.status = 'paid'"]
    params = []
    if branch_id:
        conditions.append("t.branch_id = %s")
        params.append(branch_id)
    if date_from:
        conditions.append("DATE(t.created_at) >= %s")
        params.append(date_from)
    if date_to:
        conditions.append("DATE(t.created_at) <= %s")
        params.append(date_to)
    where = " AND ".join(conditions)

    if report_type == "sales":
        rows = fetch_all(
            f"""SELECT DATE(t.created_at) as date, t.doc_key as invoice,
                       t.customer_name, pm.name as payment_method,
                       t.subtotal, t.discount, t.tax, t.total
                FROM pos_transaction t
                LEFT JOIN payment_method pm ON t.payment_method_id = pm.id
                WHERE {where}
                ORDER BY t.created_at DESC""",
            tuple(params),
        )
        headers = ["date", "invoice", "customer_name", "payment_method", "subtotal", "discount", "tax", "total"]
    elif report_type == "treatment":
        rows = fetch_all(
            f"""SELECT DATE(t.created_at) as date, ti.item_name as treatment,
                       ti.qty, ti.unit_price, ti.total as revenue,
                       t.customer_name
                FROM pos_transaction_item ti
                JOIN pos_transaction t ON ti.transaction_id = t.id
                WHERE {where} AND ti.item_type = 'treatment'
                ORDER BY t.created_at DESC""",
            tuple(params),
        )
        headers = ["date", "treatment", "qty", "unit_price", "revenue", "customer_name"]
    elif report_type == "payment":
        rows = fetch_all(
            f"""SELECT pm.name as payment_method, COUNT(*) as count,
                       SUM(t.total) as total
                FROM pos_transaction t
                LEFT JOIN payment_method pm ON t.payment_method_id = pm.id
                WHERE {where}
                GROUP BY pm.name ORDER BY total DESC""",
            tuple(params),
        )
        headers = ["payment_method", "count", "total"]
    elif report_type == "therapist":
        rows = fetch_all(
            """SELECT u.full_name as therapist_name, COUNT(*) as sessions,
                      COALESCE(SUM(ti.total), 0) as revenue
               FROM treatment_record tr
               JOIN app_user u ON tr.therapist_id = u.id
               LEFT JOIN pos_transaction_item ti ON tr.transaction_id = ti.transaction_id
               WHERE tr.status = 'completed'
               GROUP BY u.full_name ORDER BY revenue DESC""",
            (),
        )
        headers = ["therapist_name", "sessions", "revenue"]
    else:
        rows = []
        headers = []

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        writer.writerow([row.get(h, "") for h in headers])
    output.seek(0)

    filename = f"report_{report_type}_{date_from or 'all'}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Excel Export ──────────────────────────────────────────────────

@router.get("/export/excel")
def export_report_excel(
    report_type: str = "sales",
    date_from: str = "",
    date_to: str = "",
    branch_id: str = "",
):
    """Export report data as Excel file (.xlsx)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    conditions = ["t.status = 'paid'"]
    params = []
    if branch_id:
        conditions.append("t.branch_id = %s")
        params.append(branch_id)
    if date_from:
        conditions.append("DATE(t.created_at) >= %s")
        params.append(date_from)
    if date_to:
        conditions.append("DATE(t.created_at) <= %s")
        params.append(date_to)
    where = " AND ".join(conditions)

    # Get data based on report type
    if report_type == "sales":
        rows = fetch_all(
            f"""SELECT DATE(t.created_at) as date, t.doc_key as invoice,
                       t.customer_name, pm.name as payment_method,
                       t.subtotal, t.discount, t.tax, t.total
                FROM pos_transaction t
                LEFT JOIN payment_method pm ON t.payment_method_id = pm.id
                WHERE {where}
                ORDER BY t.created_at DESC""",
            tuple(params),
        )
        headers = ["Date", "Invoice", "Customer", "Payment Method", "Subtotal", "Discount", "Tax", "Total"]
        header_keys = ["date", "invoice", "customer_name", "payment_method", "subtotal", "discount", "tax", "total"]
    elif report_type == "treatment":
        rows = fetch_all(
            f"""SELECT DATE(t.created_at) as date, ti.item_name as treatment,
                       ti.qty, ti.unit_price, ti.total as revenue,
                       t.customer_name
                FROM pos_transaction_item ti
                JOIN pos_transaction t ON ti.transaction_id = t.id
                WHERE {where} AND ti.item_type = 'treatment'
                ORDER BY t.created_at DESC""",
            tuple(params),
        )
        headers = ["Date", "Treatment", "Qty", "Unit Price", "Revenue", "Customer"]
        header_keys = ["date", "treatment", "qty", "unit_price", "revenue", "customer_name"]
    elif report_type == "payment":
        rows = fetch_all(
            f"""SELECT pm.name as payment_method, COUNT(*) as count,
                       SUM(t.total) as total
                FROM pos_transaction t
                LEFT JOIN payment_method pm ON t.payment_method_id = pm.id
                WHERE {where}
                GROUP BY pm.name ORDER BY total DESC""",
            tuple(params),
        )
        headers = ["Payment Method", "Transactions", "Total"]
        header_keys = ["payment_method", "count", "total"]
    elif report_type == "therapist":
        rows = fetch_all(
            """SELECT u.full_name as therapist_name, COUNT(*) as sessions,
                      COALESCE(SUM(ti.total), 0) as revenue
               FROM treatment_record tr
               JOIN app_user u ON tr.therapist_id = u.id
               LEFT JOIN pos_transaction_item ti ON tr.transaction_id = ti.transaction_id
               WHERE tr.status = 'completed'
               GROUP BY u.full_name ORDER BY revenue DESC""",
            (),
        )
        headers = ["Therapist", "Sessions", "Revenue"]
        header_keys = ["therapist_name", "sessions", "revenue"]
    elif report_type == "shift":
        rows = fetch_all(
            """SELECT s.shift_code, s.staff_name, b.code as branch,
                      s.opening_cash, s.closing_cash, s.total_sales,
                      s.transaction_count, s.status,
                      s.opened_at, s.closed_at
               FROM pos_cashier_shift s
               LEFT JOIN branch b ON s.branch_id = b.id
               ORDER BY s.opened_at DESC LIMIT 100""",
            (),
        )
        headers = ["Shift Code", "Staff", "Branch", "Opening Cash", "Closing Cash", 
                   "Total Sales", "Transactions", "Status", "Opened", "Closed"]
        header_keys = ["shift_code", "staff_name", "branch", "opening_cash", "closing_cash",
                      "total_sales", "transaction_count", "status", "opened_at", "closed_at"]
    else:
        rows = []
        headers = []
        header_keys = []

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Report {report_type.title()}"

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="C9A96E", end_color="C9A96E", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Write title
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"Beauty & Shine - {report_type.title()} Report"
    title_cell.font = Font(bold=True, size=14, color="C9A96E")
    title_cell.alignment = Alignment(horizontal="center")

    # Write date range
    ws.merge_cells("A2:E2")
    date_cell = ws["A2"]
    date_cell.value = f"Period: {date_from or 'All'} to {date_to or 'All'}"
    date_cell.alignment = Alignment(horizontal="center")

    # Write headers (row 4)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for row_idx, row_data in enumerate(rows, 5):
        for col_idx, key in enumerate(header_keys, 1):
            value = row_data.get(key, "")
            # Convert datetime objects to string
            if hasattr(value, "isoformat"):
                value = value.isoformat()
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            # Format currency columns
            if key in ["subtotal", "discount", "tax", "total", "revenue", "unit_price",
                       "opening_cash", "closing_cash", "total_sales"]:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")

    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, 1):
        max_length = len(str(header)) + 2
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)) + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length, 30)

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"report_{report_type}_{date_from or 'all'}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Shift Report ──────────────────────────────────────────────────

@router.get("/shift/summary")
def get_shift_summary(branch_id: str = "", date_from: str = "", date_to: str = ""):
    """Get shift summary report."""
    conditions = []
    params = []
    if branch_id:
        conditions.append("s.branch_id = %s")
        params.append(branch_id)
    if date_from:
        conditions.append("DATE(s.opened_at) >= %s")
        params.append(date_from)
    if date_to:
        conditions.append("DATE(s.opened_at) <= %s")
        params.append(date_to)
    where = "WHERE " + " AND ".join(conditions) if conditions else ""

    rows = fetch_all(
        f"""SELECT s.shift_code, s.staff_name, b.code as branch,
                   s.opening_cash, s.closing_cash, s.total_sales,
                   s.total_cash_sales, s.total_qris_sales, s.total_transfer_sales,
                   s.transaction_count, s.variance, s.status,
                   s.opened_at, s.closed_at
            FROM pos_cashier_shift s
            LEFT JOIN branch b ON s.branch_id = b.id
            {where}
            ORDER BY s.opened_at DESC""",
        tuple(params),
    )
    return {"items": rows}


@router.get("/shift/staff-performance")
def get_shift_staff_performance(branch_id: str = "", date_from: str = "", date_to: str = ""):
    """Get staff performance by shift."""
    conditions = ["s.status = 'closed'"]
    params = []
    if branch_id:
        conditions.append("s.branch_id = %s")
        params.append(branch_id)
    if date_from:
        conditions.append("DATE(s.opened_at) >= %s")
        params.append(date_from)
    if date_to:
        conditions.append("DATE(s.opened_at) <= %s")
        params.append(date_to)
    where = " AND ".join(conditions)

    rows = fetch_all(
        f"""SELECT s.staff_name, s.staff_id,
                   COUNT(*) as total_shifts,
                   SUM(s.transaction_count) as total_transactions,
                   SUM(s.total_sales) as total_sales,
                   AVG(s.total_sales) as avg_sales_per_shift,
                   SUM(s.variance) as total_variance
            FROM pos_cashier_shift s
            WHERE {where}
            GROUP BY s.staff_name, s.staff_id
            ORDER BY total_sales DESC""",
        tuple(params),
    )
    return {"items": rows}


# ── Staff Commission ──────────────────────────────────────────────
# ── Staff Commission ──────────────────────────────────────────────

@router.get("/commission/summary")
def get_commission_summary(date_from: str = "", date_to: str = "", therapist_id: str = "", branch_id: str = ""):
    """Get commission summary by therapist."""
    conditions = []
    params = []
    if date_from:
        conditions.append("DATE(sc.created_at) >= %s")
        params.append(date_from)
    if date_to:
        conditions.append("DATE(sc.created_at) <= %s")
        params.append(date_to)
    if therapist_id:
        conditions.append("sc.therapist_id = %s")
        params.append(therapist_id)
    if branch_id:
        conditions.append("pt.branch_id = %s")
        params.append(branch_id)
    where = "WHERE " + " AND ".join(conditions) if conditions else ""

    rows = fetch_all(
        f"""SELECT u.full_name as therapist_name, sc.therapist_id,
                   COUNT(*) as total_commissions,
                   COALESCE(SUM(sc.commission_amount), 0) as total_amount,
                   COALESCE(SUM(CASE WHEN sc.status = 'pending' THEN sc.commission_amount ELSE 0 END), 0) as pending_amount,
                   COALESCE(SUM(CASE WHEN sc.status = 'paid' THEN sc.commission_amount ELSE 0 END), 0) as paid_amount
            FROM staff_commission sc
            JOIN app_user u ON sc.therapist_id = u.id
            LEFT JOIN pos_transaction pt ON sc.transaction_id = pt.id
            {where}
            GROUP BY u.full_name, sc.therapist_id
            ORDER BY total_amount DESC""",
        tuple(params),
    )
    return {"items": rows}

@router.get("/commission/detail")
def get_commission_detail(date_from: str = "", date_to: str = "", therapist_id: str = "", status: str = "", branch_id: str = ""):
    """Get detailed commission records."""
    conditions = []
    params = []
    if date_from:
        conditions.append("DATE(sc.created_at) >= %s")
        params.append(date_from)
    if date_to:
        conditions.append("DATE(sc.created_at) <= %s")
        params.append(date_to)
    if therapist_id:
        conditions.append("sc.therapist_id = %s")
        params.append(therapist_id)
    if status:
        conditions.append("sc.status = %s")
        params.append(status)
    if branch_id:
        conditions.append("pt.branch_id = %s")
        params.append(branch_id)
    where = "WHERE " + " AND ".join(conditions) if conditions else ""

    rows = fetch_all(
        f"""SELECT sc.*, u.full_name as therapist_name
            FROM staff_commission sc
            JOIN app_user u ON sc.therapist_id = u.id
            LEFT JOIN pos_transaction pt ON sc.transaction_id = pt.id
            {where}
            ORDER BY sc.created_at DESC LIMIT 200""",
        tuple(params),
    )
    return {"items": rows}

class CommissionPayRequest(BaseModel):
    commission_ids: list[str]

@router.post("/commission/mark-paid")
def mark_commissions_paid(req: CommissionPayRequest):
    """Mark selected commissions as paid."""
    for cid in req.commission_ids:
        execute("UPDATE staff_commission SET status='paid', paid_at=NOW() WHERE id=%s", (cid,))
    return {"message": f"{len(req.commission_ids)} commissions marked as paid"}

@router.post("/commission/generate")
def generate_commissions(date_from: str = "", date_to: str = "", branch_id: str = ""):
    """Auto-generate commissions from completed treatment records."""
    conditions = ["tr.status = 'completed'"]
    params = []
    if date_from:
        conditions.append("DATE(tr.start_time) >= %s")
        params.append(date_from)
    if date_to:
        conditions.append("DATE(tr.start_time) <= %s")
        params.append(date_to)
    if branch_id:
        conditions.append("pt.branch_id = %s")
        params.append(branch_id)
    # Only unprocessed records (no existing commission)
    conditions.append("tr.id NOT IN (SELECT treatment_record_id FROM staff_commission WHERE treatment_record_id IS NOT NULL)")
    where = " AND ".join(conditions)

    records = fetch_all(
        f"""SELECT tr.id as record_id, tr.therapist_id, tr.transaction_id,
                   t.name as treatment_name, t.price as treatment_price,
                   COALESCE(t.commission_rate, 10.00) as commission_rate
            FROM treatment_record tr
            JOIN treatment t ON tr.treatment_id = t.id
            LEFT JOIN pos_transaction pt ON tr.transaction_id = pt.id
            WHERE {where}""",
        tuple(params),
    )

    generated = 0
    for rec in records:
        price = float(rec["treatment_price"] or 0)
        rate = float(rec["commission_rate"] or 10.0)
        amount = price * rate / 100.0
        execute(
            """INSERT INTO staff_commission
               (therapist_id, treatment_record_id, transaction_id, treatment_name, treatment_price, commission_rate, commission_amount)
               VALUES (%s, %s, %s, %s, %s, %s, %s)""",
            (rec["therapist_id"], rec["record_id"], rec["transaction_id"],
             rec["treatment_name"], price, rate, amount),
        )
        generated += 1

    return {"generated": generated, "message": f"{generated} commissions generated"}
