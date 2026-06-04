"""Master Data Router - Beauty & Shine ERP."""
from __future__ import annotations
import hashlib
import json
import io
import csv
from fastapi import APIRouter, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List, Any
from pos_erp.db import fetch_all, fetch_one, execute, execute_returning

router = APIRouter(prefix="/api/master", tags=["Master"])

# ── Generic CRUD helpers ──────────────────────────────────────────

class BaseRequest(BaseModel):
    """Base for create/update requests."""
    pass

def list_items(table: str, search_field: str = "name", q: str = "", offset: int = 0, limit: int = 50, extra_where: str = "", params: tuple = ()):
    where = "WHERE is_active = true"
    p = list(params)
    if q:
        where += f" AND {search_field} ILIKE %s"
        p.append(f"%{q}%")
    if extra_where:
        where += f" AND {extra_where}"
    rows = fetch_all(f"SELECT * FROM {table} {where} ORDER BY name NULLS LAST LIMIT %s OFFSET %s", (*p, limit, offset))
    count = fetch_one(f"SELECT count(*) as total FROM {table} {where}", tuple(p))
    return {"items": rows, "total": count["total"] if count else 0}

# ── Module registry for bulk-upload / template ───────────────────

MODULE_CONFIG = {
    "product-category": {
        "table": "product_category",
        "fields": {"name": "str", "coa_id": "str"},
        "required": ["name"],
    },
    "product-subcategory": {
        "table": "product_subcategory",
        "fields": {"category_id": "str", "name": "str"},
        "required": ["category_id", "name"],
    },
    "product": {
        "table": "product",
        "fields": {"sku": "str", "name": "str", "category_id": "str", "subcategory_id": "str", "unit": "str"},
        "required": ["sku", "name"],
        "defaults": {"unit": "pcs"},
    },
    "treatment-category": {
        "table": "treatment_category",
        "fields": {"name": "str", "coa_id": "str"},
        "required": ["name"],
    },
    "treatment-subcategory": {
        "table": "treatment_subcategory",
        "fields": {"category_id": "str", "name": "str"},
        "required": ["category_id", "name"],
    },
    "treatment": {
        "table": "treatment",
        "fields": {"name": "str", "category_id": "str", "duration_minutes": "int", "price": "float", "description": "str"},
        "required": ["name"],
        "defaults": {"duration_minutes": 60, "price": 0},
    },
    "bed-section": {
        "table": "bed_section",
        "fields": {"branch_id": "str", "name": "str"},
        "required": ["branch_id", "name"],
    },
    "bed": {
        "table": "bed",
        "fields": {"section_id": "str", "name": "str", "status": "str"},
        "required": ["section_id", "name"],
        "defaults": {"status": "available"},
    },
    "coa": {
        "table": "chart_of_account",
        "fields": {"account_code": "str", "account_name": "str", "account_type": "str", "parent_code": "str", "level": "int"},
        "required": ["account_code", "account_name", "account_type"],
        "defaults": {"level": 1},
    },
    "payment-method": {
        "table": "payment_method",
        "fields": {"name": "str", "type": "str"},
        "required": ["name", "type"],
    },
    "voucher": {
        "table": "voucher",
        "fields": {"code": "str", "type": "str", "value": "float", "min_purchase": "float", "valid_from": "str", "valid_until": "str", "usage_limit": "int"},
        "required": ["code"],
        "defaults": {"type": "percentage", "value": 0, "min_purchase": 0, "usage_limit": 1},
    },
    "promotion": {
        "table": "promotion",
        "fields": {"name": "str", "type": "str", "value": "float", "applicable_to": "str", "valid_from": "str", "valid_until": "str"},
        "required": ["name"],
        "defaults": {"type": "percentage", "value": 0},
    },
    "branch": {
        "table": "branch",
        "fields": {"code": "str", "name": "str", "address": "str", "phone": "str"},
        "required": ["code", "name"],
    },
    "customer": {
        "table": "customer",
        "fields": {"name": "str", "phone": "str", "email": "str", "notes": "str"},
        "required": ["name"],
    },
    "department": {
        "table": "department",
        "fields": {"branch_id": "str", "name": "str"},
        "required": ["branch_id", "name"],
    },
    "user": {
        "table": "app_user",
        "fields": {"user_code": "str", "username": "str", "full_name": "str", "email": "str", "phone": "str", "branch_id": "str", "department_id": "str", "position_id": "str"},
        "required": ["username"],
    },
    "user-role": {
        "table": "user_role",
        "fields": {"user_id": "str", "role_name": "str", "branch_id": "str"},
        "required": ["user_id", "role_name"],
    },
    "role-permission": {
        "table": "role_permission",
        "fields": {"role_name": "str", "module": "str", "can_create": "bool", "can_read": "bool", "can_update": "bool", "can_delete": "bool", "can_approve": "bool"},
        "required": ["role_name", "module"],
        "defaults": {"can_create": False, "can_read": True, "can_update": False, "can_delete": False, "can_approve": False},
    },
    "approval-flow": {
        "table": "approval_flow",
        "fields": {"module": "str", "branch_id": "str", "min_amount": "float", "max_amount": "float", "approver_role": "str", "sequence": "int"},
        "required": ["module", "branch_id", "approver_role"],
        "defaults": {"min_amount": 0, "sequence": 1},
    },
    "financial-period": {
        "table": "financial_period",
        "fields": {"branch_id": "str", "year": "int", "month": "int", "status": "str"},
        "required": ["branch_id", "year", "month"],
        "defaults": {"status": "open"},
    },
    "account-mapping": {
        "table": "account_mapping",
        "fields": {"module": "str", "transaction_type": "str", "debit_account": "str", "credit_account": "str", "description": "str"},
        "required": ["module", "transaction_type", "debit_account", "credit_account"],
    },
    "cost-center": {
        "table": "cost_center",
        "fields": {"branch_id": "str", "name": "str", "code": "str"},
        "required": ["branch_id", "name"],
    },
    "tax-purpose": {
        "table": "tax_purpose",
        "fields": {"name": "str", "rate": "float"},
        "required": ["name"],
        "defaults": {"rate": 0},
    },
    "cancel-reason": {
        "table": "cancel_reason",
        "fields": {"module": "str", "reason": "str"},
        "required": ["module", "reason"],
    },
    "product-supplier": {
        "table": "product_supplier",
        "fields": {"product_id": "str", "supplier_name": "str", "supplier_code": "str", "lead_time_days": "int", "min_order_qty": "int"},
        "required": ["product_id", "supplier_name"],
        "defaults": {"lead_time_days": 0, "min_order_qty": 1},
    },
    "product-batch": {
        "table": "product_batch",
        "fields": {"product_id": "str", "branch_id": "str", "batch_no": "str", "expiry_date": "str", "qty": "float", "cost_per_unit": "float"},
        "required": ["product_id", "branch_id", "batch_no"],
        "defaults": {"qty": 0, "cost_per_unit": 0},
    },
    "treatment-package": {
        "table": "treatment_package",
        "fields": {"name": "str", "description": "str", "total_sessions": "int", "price": "float"},
        "required": ["name"],
        "defaults": {"total_sessions": 1, "price": 0},
    },
}


@router.post("/{module}/bulk-upload")
async def bulk_upload(module: str, request: Request):
    """Bulk insert records for a module. Expects JSON body: {"items": [...]}"""
    if module not in MODULE_CONFIG:
        raise HTTPException(status_code=404, detail=f"Module '{module}' not found")
    cfg = MODULE_CONFIG[module]
    table = cfg["table"]
    field_defs = cfg["fields"]
    required = cfg.get("required", [])
    defaults = cfg.get("defaults", {})

    body = await request.json()
    items = body.get("items", [])
    if not items:
        raise HTTPException(status_code=400, detail="No items provided")

    columns = list(field_defs.keys())
    inserted = 0
    errors = []

    for idx, item in enumerate(items):
        # Check required fields
        missing = [f for f in required if f not in item or item[f] is None]
        if missing:
            errors.append({"index": idx, "error": f"Missing required fields: {', '.join(missing)}"})
            continue

        # Build values with defaults
        vals = []
        for col in columns:
            val = item.get(col, defaults.get(col))
            # Hash password for user module
            if module == "user" and col == "full_name":
                pass  # no special handling needed
            vals.append(val)

        placeholders = ",".join(["%s"] * len(columns))
        col_str = ",".join(columns)
        try:
            execute_returning(f"INSERT INTO {table} ({col_str}) VALUES ({placeholders}) RETURNING *", tuple(vals))
            inserted += 1
        except Exception as e:
            errors.append({"index": idx, "error": str(e)})

    return {"inserted": inserted, "errors": errors, "total": len(items)}


@router.get("/{module}/template")
def download_template(module: str):
    """Get a template for bulk upload. Returns column definitions and a sample row."""
    if module not in MODULE_CONFIG:
        raise HTTPException(status_code=404, detail=f"Module '{module}' not found")
    cfg = MODULE_CONFIG[module]
    field_defs = cfg["fields"]
    required = cfg.get("required", [])
    defaults = cfg.get("defaults", {})

    columns = []
    sample = {}
    for col, ftype in field_defs.items():
        columns.append({"name": col, "type": ftype, "required": col in required, "default": defaults.get(col)})
        if col in defaults:
            sample[col] = defaults[col]
        elif ftype == "str":
            sample[col] = f"sample_{col}"
        elif ftype == "int":
            sample[col] = 0
        elif ftype == "float":
            sample[col] = 0.0
        elif ftype == "bool":
            sample[col] = False

    return {"module": module, "columns": columns, "sample_row": sample}


# ── Product Category ──────────────────────────────────────────────

@router.get("/product-category")
def list_product_categories(q: str = "", offset: int = 0, limit: int = 50):
    """List product categories with COA join."""
    where = "WHERE pc.is_active = true"
    params = []
    if q:
        where += " AND pc.name ILIKE %s"
        params.append(f"%{q}%")
    rows = fetch_all(f"""
        SELECT pc.id, pc.name, pc.is_active, pc.coa_id,
               coa.account_code AS coa_code, coa.account_name AS coa_name
        FROM product_category pc
        LEFT JOIN chart_of_account coa ON coa.id = pc.coa_id
        {where}
        ORDER BY pc.name NULLS LAST LIMIT %s OFFSET %s
    """, (*params, limit, offset))
    count = fetch_one(f"SELECT count(*) as total FROM product_category pc {where}", tuple(params))
    return {"items": rows, "total": count["total"] if count else 0}

class ProductCategoryReq(BaseModel):
    name: str
    coa_id: Optional[str] = None

@router.post("/product-category")
def create_product_category(req: ProductCategoryReq):
    return execute_returning(
        "INSERT INTO product_category (name, coa_id) VALUES (%s, %s) RETURNING *",
        (req.name, req.coa_id or None)
    )

@router.put("/product-category/{id}")
def update_product_category(id: str, req: ProductCategoryReq):
    return execute_returning(
        "UPDATE product_category SET name=%s, coa_id=%s WHERE id=%s RETURNING *",
        (req.name, req.coa_id or None, id)
    )

@router.delete("/product-category/{id}")
def delete_product_category(id: str):
    execute("UPDATE product_category SET is_active=false WHERE id=%s", (id,))
    return {"status": "deleted"}

# ── Product Subcategory ───────────────────────────────────────────

@router.get("/product-subcategory")
def list_product_subcategories(category_id: str = "", q: str = "", offset: int = 0, limit: int = 50):
    extra = "category_id = %s" if category_id else ""
    params = (category_id,) if category_id else ()
    return list_items("product_subcategory", q=q, offset=offset, limit=limit, extra_where=extra, params=params)

class ProductSubcategoryReq(BaseModel):
    category_id: str
    name: str

@router.post("/product-subcategory")
def create_product_subcategory(req: ProductSubcategoryReq):
    return execute_returning("INSERT INTO product_subcategory (category_id, name) VALUES (%s,%s) RETURNING *", (req.category_id, req.name))

@router.put("/product-subcategory/{id}")
def update_product_subcategory(id: str, req: ProductSubcategoryReq):
    return execute_returning("UPDATE product_subcategory SET category_id=%s, name=%s WHERE id=%s RETURNING *", (req.category_id, req.name, id))

@router.delete("/product-subcategory/{id}")
def delete_product_subcategory(id: str):
    execute("UPDATE product_subcategory SET is_active=false WHERE id=%s", (id,))
    return {"status": "deleted"}

# ── Product ───────────────────────────────────────────────────────

@router.get("/product")
def list_products(q: str = "", category_id: str = "", offset: int = 0, limit: int = 50):
    """List products with subcategory name, BOM name, and total stock qty."""
    where = "WHERE p.is_active = true"
    params: list = []
    if q:
        where += " AND p.name ILIKE %s"
        params.append(f"%{q}%")
    if category_id:
        where += " AND p.category_id = %s"
        params.append(category_id)

    rows = fetch_all(f"""
        SELECT p.*,
               ps.name AS subcategory_name,
               pc.name AS category_name,
               bh.name AS bom_name,
               COALESCE(SUM(pb.qty), 0) AS qty
        FROM product p
        LEFT JOIN product_subcategory ps ON ps.id = p.subcategory_id
        LEFT JOIN product_category pc ON pc.id = p.category_id
        LEFT JOIN bom_header bh ON bh.product_id = p.id AND bh.is_active = true
        LEFT JOIN product_batch pb ON pb.product_id = p.id
        {where}
        GROUP BY p.id, ps.name, pc.name, bh.name
        ORDER BY p.name NULLS LAST
        LIMIT %s OFFSET %s
    """, (*params, limit, offset))

    count = fetch_one(f"SELECT count(*) as total FROM product p {where}", tuple(params))
    return {"items": rows, "total": count["total"] if count else 0}


@router.get("/product/export")
def export_products():
    """Export all active products as JSON for download."""
    rows = fetch_all("""
        SELECT p.id AS product_id, p.sku, p.name,
               pc.name AS category, ps.name AS subcategory,
               bh.name AS bom_name, p.unit, p.barcode,
               COALESCE(SUM(pb.qty), 0) AS qty,
               p.receipt_tolerance, p.status_uom, p.min_stock_threshold
        FROM product p
        LEFT JOIN product_subcategory ps ON ps.id = p.subcategory_id
        LEFT JOIN product_category pc ON pc.id = p.category_id
        LEFT JOIN bom_header bh ON bh.product_id = p.id AND bh.is_active = true
        LEFT JOIN product_batch pb ON pb.product_id = p.id
        WHERE p.is_active = true
        GROUP BY p.id, pc.name, ps.name, bh.name
        ORDER BY p.name
    """)
    return {"items": rows}


def _get_export_rows():
    """Shared query for all export formats."""
    return fetch_all("""
        SELECT p.id AS product_id, p.sku, p.name,
               pc.name AS category, ps.name AS subcategory,
               bh.name AS bom_name, p.unit, p.barcode,
               COALESCE(SUM(pb.qty), 0) AS qty,
               p.receipt_tolerance, p.status_uom, p.min_stock_threshold
        FROM product p
        LEFT JOIN product_subcategory ps ON ps.id = p.subcategory_id
        LEFT JOIN product_category pc ON pc.id = p.category_id
        LEFT JOIN bom_header bh ON bh.product_id = p.id AND bh.is_active = true
        LEFT JOIN product_batch pb ON pb.product_id = p.id
        WHERE p.is_active = true
        GROUP BY p.id, pc.name, ps.name, bh.name
        ORDER BY p.name
    """)


@router.get("/product/export/csv")
def export_products_csv():
    """Export products as CSV file."""
    rows = _get_export_rows()
    headers = ['Product ID', 'SKU', 'Name', 'Category', 'Sub Category', 'BOM Name', 'Unit', 'Barcode', 'Qty', 'Receipt Tolerance (%)', 'Status UOM', 'Min Stock Threshold']
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for r in rows:
        writer.writerow([r.get('product_id',''), r.get('sku',''), r.get('name',''), r.get('category',''), r.get('subcategory',''),
                         r.get('bom_name',''), r.get('unit','pcs'), r.get('barcode',''), r.get('qty',0),
                         r.get('receipt_tolerance',0), r.get('status_uom','active'), r.get('min_stock_threshold',10)])
    output.seek(0)
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv",
                             headers={"Content-Disposition": "attachment; filename=products-export.csv"})


@router.get("/product/export/xlsx")
def export_products_xlsx():
    """Export products as XLSX file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    rows = _get_export_rows()
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    headers = ['Product ID', 'SKU', 'Name', 'Category', 'Sub Category', 'BOM Name', 'Unit', 'Barcode', 'Qty', 'Receipt Tolerance (%)', 'Status UOM', 'Min Stock Threshold']
    header_fill = PatternFill(start_color="C9A96E", end_color="C9A96E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    for row_idx, r in enumerate(rows, 2):
        vals = [str(r.get('product_id','')), r.get('sku',''), r.get('name',''), r.get('category',''), r.get('subcategory',''),
                r.get('bom_name',''), r.get('unit','pcs'), r.get('barcode',''), float(r.get('qty',0)),
                float(r.get('receipt_tolerance',0)), r.get('status_uom','active'), float(r.get('min_stock_threshold',10))]
        for col_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.border = thin_border
    for col in range(1, len(headers)+1):
        ws.column_dimensions[chr(64+col) if col <= 26 else 'A'].width = 16
    ws.column_dimensions['C'].width = 30  # Name
    ws.column_dimensions['A'].width = 38  # Product ID
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=products-export.xlsx"})


@router.get("/product/export/pdf")
def export_products_pdf():
    """Export products as PDF file."""
    from fpdf import FPDF
    rows = _get_export_rows()
    headers = ['Product ID', 'SKU', 'Name', 'Category', 'Sub Cat', 'BOM', 'Unit', 'Barcode', 'Qty', 'Tol%', 'UOM', 'MinStk']
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    # Title
    pdf.set_font('Helvetica', 'B', 14)
    pdf.cell(0, 10, 'Beauty & Shine - Product Export', ln=True, align='C')
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(0, 6, f'Total: {len(rows)} products', ln=True, align='C')
    pdf.ln(4)
    # Column widths (landscape A4 = 277mm usable)
    widths = [22, 16, 48, 24, 24, 22, 12, 20, 12, 12, 14, 14]
    # Header
    pdf.set_fill_color(201, 169, 110)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 7)
    for i, h in enumerate(headers):
        pdf.cell(widths[i], 7, h, border=1, fill=True, align='C')
    pdf.ln()
    # Data rows
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Helvetica', '', 6.5)
    for r in rows:
        vals = [
            str(r.get('product_id',''))[:8]+'...',
            r.get('sku','') or '',
            r.get('name','') or '',
            r.get('category','') or '-',
            r.get('subcategory','') or '-',
            r.get('bom_name','') or '-',
            r.get('unit','pcs') or 'pcs',
            r.get('barcode','') or '-',
            str(r.get('qty',0)),
            str(r.get('receipt_tolerance',0)),
            r.get('status_uom','active') or 'active',
            str(r.get('min_stock_threshold',10)),
        ]
        for i, v in enumerate(vals):
            pdf.cell(widths[i], 6, v[:30], border=1)
        pdf.ln()
    output = io.BytesIO()
    output.write(pdf.output())
    output.seek(0)
    return StreamingResponse(output, media_type="application/pdf",
                             headers={"Content-Disposition": "attachment; filename=products-export.pdf"})

class ProductReq(BaseModel):
    sku: str
    name: str
    category_id: Optional[str] = None
    subcategory_id: Optional[str] = None
    unit: str = "pcs"
    barcode: Optional[str] = None
    receipt_tolerance: Optional[float] = 0
    status_uom: Optional[str] = "active"
    min_stock_threshold: Optional[float] = 10
    description: Optional[str] = None

class ProductUpdateReq(BaseModel):
    sku: Optional[str] = None
    name: Optional[str] = None
    category_id: Optional[str] = None
    subcategory_id: Optional[str] = None
    unit: Optional[str] = None
    barcode: Optional[str] = None
    receipt_tolerance: Optional[float] = None
    status_uom: Optional[str] = None
    min_stock_threshold: Optional[float] = None
    description: Optional[str] = None

@router.post("/product")
def create_product(req: ProductReq):
    return execute_returning(
        "INSERT INTO product (sku, name, category_id, subcategory_id, unit, barcode, receipt_tolerance, status_uom, min_stock_threshold) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *",
        (req.sku, req.name, req.category_id, req.subcategory_id, req.unit, req.barcode, req.receipt_tolerance or 0, req.status_uom or "active", req.min_stock_threshold or 10)
    )

@router.put("/product/{id}")
def update_product(id: str, req: ProductUpdateReq):
    fields = {}
    if req.sku is not None:
        fields["sku"] = req.sku
    if req.name is not None:
        fields["name"] = req.name
    if req.category_id is not None:
        fields["category_id"] = req.category_id
    if req.subcategory_id is not None:
        fields["subcategory_id"] = req.subcategory_id
    if req.unit is not None:
        fields["unit"] = req.unit
    if req.barcode is not None:
        fields["barcode"] = req.barcode
    if req.receipt_tolerance is not None:
        fields["receipt_tolerance"] = req.receipt_tolerance
    if req.status_uom is not None:
        fields["status_uom"] = req.status_uom
    if req.min_stock_threshold is not None:
        fields["min_stock_threshold"] = req.min_stock_threshold
    if not fields:
        raise HTTPException(status_code=400, detail="No fields to update")
    set_parts = [f"{k}=%s" for k in fields]
    vals = list(fields.values())
    vals.append(id)
    return execute_returning(
        f"UPDATE product SET {', '.join(set_parts)} WHERE id=%s RETURNING *",
        tuple(vals)
    )

@router.delete("/product/{id}")
def delete_product(id: str):
    execute("UPDATE product SET is_active=false WHERE id=%s", (id,))
    return {"status": "deleted"}

# ── Treatment Category ────────────────────────────────────────────

@router.get("/treatment-category")
def list_treatment_categories(q: str = "", offset: int = 0, limit: int = 50):
    """List treatment categories with COA join."""
    where = "WHERE tc.is_active = true"
    params = []
    if q:
        where += " AND tc.name ILIKE %s"
        params.append(f"%{q}%")
    rows = fetch_all(f"""
        SELECT tc.id, tc.name, tc.is_active, tc.coa_id,
               coa.account_code AS coa_code, coa.account_name AS coa_name
        FROM treatment_category tc
        LEFT JOIN chart_of_account coa ON coa.id = tc.coa_id
        {where}
        ORDER BY tc.name NULLS LAST LIMIT %s OFFSET %s
    """, (*params, limit, offset))
    count = fetch_one(f"SELECT count(*) as total FROM treatment_category tc {where}", tuple(params))
    return {"items": rows, "total": count["total"] if count else 0}

class TreatmentCategoryReq(BaseModel):
    name: str
    coa_id: Optional[str] = None

@router.post("/treatment-category")
def create_treatment_category(req: TreatmentCategoryReq):
    return execute_returning(
        "INSERT INTO treatment_category (name, coa_id) VALUES (%s, %s) RETURNING *",
        (req.name, req.coa_id or None)
    )

@router.put("/treatment-category/{id}")
def update_treatment_category(id: str, req: TreatmentCategoryReq):
    return execute_returning(
        "UPDATE treatment_category SET name=%s, coa_id=%s WHERE id=%s RETURNING *",
        (req.name, req.coa_id or None, id)
    )

@router.delete("/treatment-category/{id}")
def delete_treatment_category(id: str):
    execute("UPDATE treatment_category SET is_active=false WHERE id=%s", (id,))
    return {"status": "deleted"}

# ── Treatment Subcategory ────────────────────────────────────────

@router.get("/treatment-subcategory")
def list_treatment_subcategories(category_id: str = "", q: str = "", offset: int = 0, limit: int = 50):
    extra = "category_id = %s" if category_id else ""
    params = (category_id,) if category_id else ()
    return list_items("treatment_subcategory", q=q, offset=offset, limit=limit, extra_where=extra, params=params)

class TreatmentSubcategoryReq(BaseModel):
    category_id: str
    name: str

@router.post("/treatment-subcategory")
def create_treatment_subcategory(req: TreatmentSubcategoryReq):
    return execute_returning("INSERT INTO treatment_subcategory (category_id, name) VALUES (%s,%s) RETURNING *", (req.category_id, req.name))

@router.put("/treatment-subcategory/{id}")
def update_treatment_subcategory(id: str, req: TreatmentSubcategoryReq):
    return execute_returning("UPDATE treatment_subcategory SET category_id=%s, name=%s WHERE id=%s RETURNING *", (req.category_id, req.name, id))

@router.delete("/treatment-subcategory/{id}")
def delete_treatment_subcategory(id: str):
    execute("UPDATE treatment_subcategory SET is_active=false WHERE id=%s", (id,))
    return {"status": "deleted"}

# ── Treatment ─────────────────────────────────────────────────────

@router.get("/treatment")
def list_treatments(category_id: str = "", q: str = "", offset: int = 0, limit: int = 50):
    extra = "category_id = %s" if category_id else ""
    params = (category_id,) if category_id else ()
    return list_items("treatment", q=q, offset=offset, limit=limit, extra_where=extra, params=params)

class TreatmentReq(BaseModel):
    name: str
    category_id: Optional[str] = None
    duration_minutes: int = 60
    price: float = 0
    description: Optional[str] = None

@router.post("/treatment")
def create_treatment(req: TreatmentReq):
    return execute_returning(
        "INSERT INTO treatment (name, category_id, duration_minutes, price, description) VALUES (%s,%s,%s,%s,%s) RETURNING *",
        (req.name, req.category_id, req.duration_minutes, req.price, req.description)
    )

@router.put("/treatment/{id}")
def update_treatment(id: str, req: TreatmentReq):
    return execute_returning(
        "UPDATE treatment SET name=%s, category_id=%s, duration_minutes=%s, price=%s, description=%s WHERE id=%s RETURNING *",
        (req.name, req.category_id, req.duration_minutes, req.price, req.description, id)
    )

@router.delete("/treatment/{id}")
def delete_treatment(id: str):
    execute("UPDATE treatment SET is_active=false WHERE id=%s", (id,))
    return {"status": "deleted"}

# ── Bed Section ───────────────────────────────────────────────────

@router.get("/bed-section")
def list_bed_sections(branch_id: str = "", offset: int = 0, limit: int = 50):
    extra = "branch_id = %s" if branch_id else ""
    params = (branch_id,) if branch_id else ()
    return list_items("bed_section", q="", offset=offset, limit=limit, extra_where=extra, params=params)

class BedSectionReq(BaseModel):
    branch_id: str
    name: str

@router.post("/bed-section")
def create_bed_section(req: BedSectionReq):
    return execute_returning("INSERT INTO bed_section (branch_id, name) VALUES (%s,%s) RETURNING *", (req.branch_id, req.name))

@router.put("/bed-section/{id}")
def update_bed_section(id: str, req: BedSectionReq):
    return execute_returning("UPDATE bed_section SET branch_id=%s, name=%s WHERE id=%s RETURNING *", (req.branch_id, req.name, id))

# ── Bed ───────────────────────────────────────────────────────────

@router.get("/bed")
def list_beds(section_id: str = "", offset: int = 0, limit: int = 50):
    extra = "section_id = %s" if section_id else ""
    params = (section_id,) if section_id else ()
    return list_items("bed", q="", offset=offset, limit=limit, extra_where=extra, params=params)

class BedReq(BaseModel):
    section_id: str
    name: str
    status: str = "available"

@router.post("/bed")
def create_bed(req: BedReq):
    return execute_returning("INSERT INTO bed (section_id, name, status) VALUES (%s,%s,%s) RETURNING *", (req.section_id, req.name, req.status))

@router.put("/bed/{id}")
def update_bed(id: str, req: BedReq):
    return execute_returning("UPDATE bed SET section_id=%s, name=%s, status=%s WHERE id=%s RETURNING *", (req.section_id, req.name, req.status, id))

# ── Chart of Account ──────────────────────────────────────────────

@router.get("/coa")
def list_coa(q: str = "", account_type: str = "", offset: int = 0, limit: int = 100):
    conditions = ["is_active = true"]
    params = []
    if account_type:
        conditions.append("account_type = %s")
        params.append(account_type)
    if q:
        conditions.append("account_name ILIKE %s")
        params.append(f"%{q}%")
    where = " AND ".join(conditions)
    params.extend([limit, offset])
    rows = fetch_all(f"SELECT * FROM chart_of_account WHERE {where} ORDER BY account_code LIMIT %s OFFSET %s", tuple(params))
    return {"items": rows}

class COACreateReq(BaseModel):
    account_code: str
    account_name: str
    account_type: str
    parent_code: Optional[str] = None
    level: int = 1

class COAUpdateReq(BaseModel):
    account_code: Optional[str] = None
    account_name: Optional[str] = None
    account_type: Optional[str] = None
    parent_code: Optional[str] = None
    level: Optional[int] = None

@router.post("/coa")
def create_coa(req: COACreateReq):
    return execute_returning(
        "INSERT INTO chart_of_account (account_code, account_name, account_type, parent_code, level) VALUES (%s,%s,%s,%s,%s) RETURNING *",
        (req.account_code, req.account_name, req.account_type, req.parent_code, req.level)
    )

@router.put("/coa/{id}")
def update_coa(id: str, req: COAUpdateReq):
    fields = {}
    if req.account_code is not None:
        fields["account_code"] = req.account_code
    if req.account_name is not None:
        fields["account_name"] = req.account_name
    if req.account_type is not None:
        fields["account_type"] = req.account_type
    if req.parent_code is not None:
        fields["parent_code"] = req.parent_code
    if req.level is not None:
        fields["level"] = req.level
    if not fields:
        raise HTTPException(status_code=400, detail="No fields to update")
    set_parts = [f"{k}=%s" for k in fields]
    vals = list(fields.values())
    vals.append(id)
    return execute_returning(
        f"UPDATE chart_of_account SET {', '.join(set_parts)} WHERE id=%s RETURNING *",
        tuple(vals)
    )

@router.delete("/coa/{id}")
def delete_coa(id: str):
    execute("UPDATE chart_of_account SET is_active=false WHERE id=%s", (id,))
    return {"status": "deleted"}

# ── Payment Method ────────────────────────────────────────────────

@router.get("/payment-method")
def list_payment_methods():
    return {"items": fetch_all("SELECT * FROM payment_method WHERE is_active=true ORDER BY name")}

class PaymentMethodReq(BaseModel):
    name: str
    type: str

@router.post("/payment-method")
def create_payment_method(req: PaymentMethodReq):
    return execute_returning("INSERT INTO payment_method (name, type) VALUES (%s,%s) RETURNING *", (req.name, req.type))

# ── Voucher ───────────────────────────────────────────────────────

@router.get("/voucher")
def list_vouchers(q: str = "", offset: int = 0, limit: int = 50):
    """List vouchers."""
    conditions = ["is_active = true"]
    params = []
    if q:
        conditions.append("code ILIKE %s")
        params.append(f"%{q}%")
    where = " AND ".join(conditions)
    params.extend([limit, offset])
    rows = fetch_all(f"SELECT * FROM voucher WHERE {where} ORDER BY code LIMIT %s OFFSET %s", tuple(params))
    return {"items": rows}

class VoucherReq(BaseModel):
    code: str
    type: str = "percentage"
    value: float = 0
    min_purchase: float = 0
    valid_from: Optional[str] = None
    valid_until: Optional[str] = None
    usage_limit: int = 1

@router.post("/voucher")
def create_voucher(req: VoucherReq):
    return execute_returning(
        "INSERT INTO voucher (code, type, value, min_purchase, valid_from, valid_until, usage_limit) VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING *",
        (req.code, req.type, req.value, req.min_purchase, req.valid_from, req.valid_until, req.usage_limit)
    )

@router.put("/voucher/{id}")
def update_voucher(id: str, req: VoucherReq):
    return execute_returning(
        "UPDATE voucher SET code=%s, type=%s, value=%s, min_purchase=%s, valid_from=%s, valid_until=%s, usage_limit=%s WHERE id=%s RETURNING *",
        (req.code, req.type, req.value, req.min_purchase, req.valid_from, req.valid_until, req.usage_limit, id)
    )

@router.delete("/voucher/{id}")
def delete_voucher(id: str):
    execute("UPDATE voucher SET is_active=false WHERE id=%s", (id,))
    return {"status": "deleted"}

# ── Promotion ─────────────────────────────────────────────────────

@router.get("/promotion")
def list_promotions(q: str = "", offset: int = 0, limit: int = 50):
    return list_items("promotion", q=q, offset=offset, limit=limit)

class PromotionReq(BaseModel):
    name: str
    type: str = "percentage"
    value: float = 0
    applicable_to: Optional[str] = None
    valid_from: Optional[str] = None
    valid_until: Optional[str] = None

@router.post("/promotion")
def create_promotion(req: PromotionReq):
    return execute_returning(
        "INSERT INTO promotion (name, type, value, applicable_to, valid_from, valid_until) VALUES (%s,%s,%s,%s,%s,%s) RETURNING *",
        (req.name, req.type, req.value, req.applicable_to, req.valid_from, req.valid_until)
    )

@router.put("/promotion/{id}")
def update_promotion(id: str, req: PromotionReq):
    return execute_returning(
        "UPDATE promotion SET name=%s, type=%s, value=%s, applicable_to=%s, valid_from=%s, valid_until=%s WHERE id=%s RETURNING *",
        (req.name, req.type, req.value, req.applicable_to, req.valid_from, req.valid_until, id)
    )

@router.delete("/promotion/{id}")
def delete_promotion(id: str):
    execute("UPDATE promotion SET is_active=false WHERE id=%s", (id,))
    return {"status": "deleted"}

# ── Branch ────────────────────────────────────────────────────────

@router.get("/branch")
def list_branches():
    return {"items": fetch_all("SELECT * FROM branch WHERE is_active=true ORDER BY name")}

class BranchReq(BaseModel):
    code: str
    name: str
    address: Optional[str] = None
    phone: Optional[str] = None

@router.post("/branch")
def create_branch(req: BranchReq):
    return execute_returning(
        "INSERT INTO branch (code, name, address, phone) VALUES (%s,%s,%s,%s) RETURNING *",
        (req.code, req.name, req.address, req.phone)
    )

@router.put("/branch/{id}")
def update_branch(id: str, req: BranchReq):
    return execute_returning(
        "UPDATE branch SET code=%s, name=%s, address=%s, phone=%s WHERE id=%s RETURNING *",
        (req.code, req.name, req.address, req.phone, id)
    )

@router.delete("/branch/{id}")
def delete_branch(id: str):
    execute("UPDATE branch SET is_active=false WHERE id=%s", (id,))
    return {"status": "deleted"}

# ── Customer ──────────────────────────────────────────────────────

@router.get("/customer")
def list_customers(q: str = "", offset: int = 0, limit: int = 50):
    where = "WHERE is_active = true"
    params = []
    if q:
        where += " AND (name ILIKE %s OR phone ILIKE %s)"
        params.extend([f"%{q}%", f"%{q}%"])
    rows = fetch_all(
        f"SELECT * FROM customer {where} ORDER BY name LIMIT %s OFFSET %s",
        (*params, limit, offset)
    )
    count = fetch_one(f"SELECT count(*) as total FROM customer {where}", tuple(params))
    return {"items": rows, "total": count["total"] if count else 0}

class CustomerReq(BaseModel):
    name: str
    phone: Optional[str] = None
    email: Optional[str] = None
    notes: Optional[str] = None

@router.post("/customer")
def create_customer(req: CustomerReq):
    return execute_returning(
        "INSERT INTO customer (name, phone, email, notes) VALUES (%s,%s,%s,%s) RETURNING *",
        (req.name, req.phone, req.email, req.notes)
    )

@router.put("/customer/{id}")
def update_customer(id: str, req: CustomerReq):
    return execute_returning(
        "UPDATE customer SET name=%s, phone=%s, email=%s, notes=%s, updated_at=NOW() WHERE id=%s RETURNING *",
        (req.name, req.phone, req.email, req.notes, id)
    )

@router.delete("/customer/{id}")
def delete_customer(id: str):
    execute("UPDATE customer SET is_active=false WHERE id=%s", (id,))
    return {"status": "deleted"}


# ── Loyalty Points ───────────────────────────────────────────────

TIER_THRESHOLDS = [
    ("Platinum", 10000000),  # 10 juta
    ("Gold", 5000000),       # 5 juta
    ("Silver", 2000000),     # 2 juta
    ("Bronze", 0),
]

def _calculate_tier(total_spent: float) -> str:
    for tier, threshold in TIER_THRESHOLDS:
        if total_spent >= threshold:
            return tier
    return "Bronze"

@router.get("/loyalty/summary")
def get_loyalty_summary():
    """Get loyalty program summary."""
    tiers = fetch_all(
        "SELECT loyalty_tier, COUNT(*) as count, COALESCE(SUM(loyalty_points), 0) as total_points FROM customer WHERE is_active=true GROUP BY loyalty_tier ORDER BY loyalty_tier"
    )
    total_customers = fetch_one("SELECT COUNT(*) as count FROM customer WHERE is_active=true")
    total_points = fetch_one("SELECT COALESCE(SUM(loyalty_points), 0) as total FROM customer WHERE is_active=true")
    top_loyal = fetch_all(
        "SELECT name, phone, loyalty_points, loyalty_tier, total_spent FROM customer WHERE is_active=true ORDER BY loyalty_points DESC LIMIT 10"
    )
    return {
        "tiers": tiers,
        "total_customers": total_customers["count"] if total_customers else 0,
        "total_points_outstanding": total_points["total"] if total_points else 0,
        "top_loyal_customers": top_loyal,
    }

@router.get("/loyalty/customer/{customer_id}")
def get_customer_loyalty(customer_id: str):
    """Get loyalty details for a customer."""
    customer = fetch_one(
        "SELECT id, name, phone, email, loyalty_points, loyalty_tier, total_spent, last_visit_at FROM customer WHERE id=%s",
        (customer_id,)
    )
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    history = fetch_all(
        "SELECT * FROM loyalty_transaction WHERE customer_id=%s ORDER BY created_at DESC LIMIT 50",
        (customer_id,)
    )
    return {"customer": customer, "history": history}

class LoyaltyEarnRequest(BaseModel):
    points: int
    transaction_id: Optional[str] = None
    description: str = ""

@router.post("/loyalty/customer/{customer_id}/earn")
def earn_points(customer_id: str, req: LoyaltyEarnRequest):
    """Manually add loyalty points to a customer."""
    customer = fetch_one("SELECT * FROM customer WHERE id=%s", (customer_id,))
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    new_points = (customer["loyalty_points"] or 0) + req.points
    execute(
        "UPDATE customer SET loyalty_points=%s, updated_at=NOW() WHERE id=%s",
        (new_points, customer_id)
    )
    execute(
        "INSERT INTO loyalty_transaction (customer_id, transaction_id, points_change, reason, description) VALUES (%s,%s,%s,%s,%s)",
        (customer_id, req.transaction_id, req.points, "earn", req.description or f"Earned {req.points} points")
    )
    return {"loyalty_points": new_points, "message": f"+{req.points} points earned"}

class LoyaltyRedeemRequest(BaseModel):
    points: int
    description: str = ""

@router.post("/loyalty/customer/{customer_id}/redeem")
def redeem_points(customer_id: str, req: LoyaltyRedeemRequest):
    """Redeem loyalty points from a customer."""
    customer = fetch_one("SELECT * FROM customer WHERE id=%s", (customer_id,))
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    if (customer["loyalty_points"] or 0) < req.points:
        raise HTTPException(status_code=400, detail="Insufficient points")
    
    new_points = (customer["loyalty_points"] or 0) - req.points
    execute(
        "UPDATE customer SET loyalty_points=%s, updated_at=NOW() WHERE id=%s",
        (new_points, customer_id)
    )
    execute(
        "INSERT INTO loyalty_transaction (customer_id, points_change, reason, description) VALUES (%s,%s,%s,%s)",
        (customer_id, -req.points, "redeem", req.description or f"Redeemed {req.points} points")
    )
    return {"loyalty_points": new_points, "message": f"-{req.points} points redeemed"}

class LoyaltyAutoEarnRequest(BaseModel):
    customer_id: str
    transaction_id: str
    amount: float

@router.post("/loyalty/auto-earn")
def auto_earn_from_transaction(req: LoyaltyAutoEarnRequest):
    """Auto-earn points when a transaction is paid (1 point per Rp 10,000)."""
    customer = fetch_one("SELECT * FROM customer WHERE id=%s", (req.customer_id,))
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    points_earned = int(req.amount / 10000)  # 1 point per 10rb
    if points_earned <= 0:
        return {"loyalty_points": customer["loyalty_points"], "points_earned": 0}
    
    new_points = (customer["loyalty_points"] or 0) + points_earned
    new_total_spent = float(customer["total_spent"] or 0) + req.amount
    new_tier = _calculate_tier(new_total_spent)
    
    execute(
        "UPDATE customer SET loyalty_points=%s, total_spent=%s, loyalty_tier=%s, last_visit_at=NOW(), updated_at=NOW() WHERE id=%s",
        (new_points, new_total_spent, new_tier, req.customer_id)
    )
    execute(
        "INSERT INTO loyalty_transaction (customer_id, transaction_id, points_change, reason, description) VALUES (%s,%s,%s,%s,%s)",
        (req.customer_id, req.transaction_id, points_earned, "earn", f"Auto-earn from transaction ({req.amount:.0f})")
    )
    
    tier_changed = new_tier != (customer["loyalty_tier"] or "Bronze")
    return {
        "loyalty_points": new_points,
        "points_earned": points_earned,
        "tier": new_tier,
        "tier_changed": tier_changed,
        "total_spent": new_total_spent,
    }

@router.get("/loyalty/leaderboard")
def get_loyalty_leaderboard(limit: int = 20):
    """Get top loyalty customers."""
    rows = fetch_all(
        "SELECT id, name, phone, loyalty_points, loyalty_tier, total_spent, last_visit_at FROM customer WHERE is_active=true ORDER BY loyalty_points DESC LIMIT %s",
        (limit,)
    )
    return {"items": rows}


# ── Department ────────────────────────────────────────────────────

@router.get("/department")
def list_departments(branch_id: str = ""):
    if branch_id:
        return {"items": fetch_all("SELECT * FROM department WHERE is_active=true AND branch_id=%s ORDER BY name", (branch_id,))}
    return {"items": fetch_all("SELECT * FROM department WHERE is_active=true ORDER BY name")}

class DepartmentReq(BaseModel):
    branch_id: str
    name: str

@router.post("/department")
def create_department(req: DepartmentReq):
    return execute_returning("INSERT INTO department (branch_id, name) VALUES (%s,%s) RETURNING *", (req.branch_id, req.name))

@router.delete("/department/{id}")
def delete_department(id: str):
    execute("UPDATE department SET is_active=false WHERE id=%s", (id,))
    return {"status": "deleted"}

# ── User ──────────────────────────────────────────────────────────

@router.get("/user")
def list_users(branch_id: str = "", q: str = "", offset: int = 0, limit: int = 50):
    extra = "u.branch_id = %s" if branch_id else ""
    params = (branch_id,) if branch_id else ()
    where = "WHERE u.is_active = true"
    if q:
        where += " AND (u.username ILIKE %s OR u.full_name ILIKE %s)"
        params = (*params, f"%{q}%", f"%{q}%")
    if extra:
        where += f" AND {extra}"
    rows = fetch_all(
        f"SELECT u.*, b.name as branch_name, d.name as department_name FROM app_user u LEFT JOIN branch b ON u.branch_id=b.id LEFT JOIN department d ON u.department_id=d.id {where} ORDER BY u.full_name LIMIT %s OFFSET %s",
        (*params, limit, offset)
    )
    return {"items": rows}

class UserReq(BaseModel):
    user_code: Optional[str] = None
    username: str
    password: Optional[str] = None
    password_hash: Optional[str] = None
    pin: Optional[str] = None
    full_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    branch_id: Optional[str] = None
    department_id: Optional[str] = None
    position_id: Optional[str] = None

def _hash_password(password: str) -> str:
    """Simple SHA-256 password hash (matches app convention)."""
    return hashlib.sha256(password.encode()).hexdigest()

@router.post("/user")
def create_user(req: UserReq):
    pw_hash = req.password_hash
    if not pw_hash and req.password:
        pw_hash = _hash_password(req.password)
    if not pw_hash:
        raise HTTPException(status_code=400, detail="password or password_hash is required")
    return execute_returning(
        "INSERT INTO app_user (user_code, username, password_hash, pin, full_name, email, phone, branch_id, department_id, position_id) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *",
        (req.user_code, req.username, pw_hash, req.pin, req.full_name or req.username, req.email, req.phone, req.branch_id, req.department_id, req.position_id)
    )

@router.put("/user/{id}")
def update_user(id: str, req: UserReq):
    fields = {}
    if req.user_code is not None:
        fields["user_code"] = req.user_code
    if req.username is not None:
        fields["username"] = req.username
    if req.full_name is not None:
        fields["full_name"] = req.full_name
    if req.email is not None:
        fields["email"] = req.email
    if req.phone is not None:
        fields["phone"] = req.phone
    if req.branch_id is not None:
        fields["branch_id"] = req.branch_id
    if req.department_id is not None:
        fields["department_id"] = req.department_id
    if req.position_id is not None:
        fields["position_id"] = req.position_id
    if req.pin is not None:
        fields["pin"] = req.pin
    if req.password:
        fields["password_hash"] = _hash_password(req.password)
    elif req.password_hash:
        fields["password_hash"] = req.password_hash
    if not fields:
        raise HTTPException(status_code=400, detail="No fields to update")
    set_parts = [f"{k}=%s" for k in fields]
    vals = list(fields.values())
    vals.append(id)
    return execute_returning(
        f"UPDATE app_user SET {', '.join(set_parts)} WHERE id=%s RETURNING *",
        tuple(vals)
    )

@router.delete("/user/{id}")
def delete_user(id: str):
    execute("UPDATE app_user SET is_active=false WHERE id=%s", (id,))
    return {"status": "deleted"}

# ── User Role ─────────────────────────────────────────────────────

@router.get("/user-role")
def list_user_roles(user_id: str = ""):
    if user_id:
        return {"items": fetch_all("SELECT * FROM user_role WHERE is_active=true AND user_id=%s", (user_id,))}
    return {"items": fetch_all("SELECT ur.*, u.username, u.full_name FROM user_role ur JOIN app_user u ON ur.user_id=u.id WHERE ur.is_active=true")}

class UserRoleReq(BaseModel):
    user_id: str
    role_name: str
    branch_id: Optional[str] = None

@router.post("/user-role")
def create_user_role(req: UserRoleReq):
    return execute_returning("INSERT INTO user_role (user_id, role_name, branch_id) VALUES (%s,%s,%s) RETURNING *", (req.user_id, req.role_name, req.branch_id))

@router.delete("/user-role/{id}")
def delete_user_role(id: str):
    execute("UPDATE user_role SET is_active=false WHERE id=%s", (id,))
    return {"status": "deleted"}

# ── Role Permission ───────────────────────────────────────────────

@router.get("/role-permission")
def list_role_permissions(role_name: str = ""):
    if role_name:
        return {"items": fetch_all("SELECT * FROM role_permission WHERE role_name=%s", (role_name,))}
    return {"items": fetch_all("SELECT * FROM role_permission ORDER BY role_name, module")}

class RolePermissionReq(BaseModel):
    role_name: str
    module: str
    can_create: bool = False
    can_read: bool = True
    can_update: bool = False
    can_delete: bool = False
    can_approve: bool = False

@router.post("/role-permission")
def create_role_permission(req: RolePermissionReq):
    return execute_returning(
        "INSERT INTO role_permission (role_name, module, can_create, can_read, can_update, can_delete, can_approve) VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING *",
        (req.role_name, req.module, req.can_create, req.can_read, req.can_update, req.can_delete, req.can_approve)
    )

@router.delete("/role-permission/{id}")
def delete_role_permission(id: str):
    execute("DELETE FROM role_permission WHERE id=%s", (id,))
    return {"status": "deleted"}

# ── Approval Flow ─────────────────────────────────────────────────

@router.get("/approval-flow")
def list_approval_flows(branch_id: str = ""):
    if branch_id:
        return {"items": fetch_all("SELECT * FROM approval_flow WHERE is_active=true AND branch_id=%s ORDER BY sequence", (branch_id,))}
    return {"items": fetch_all("SELECT * FROM approval_flow WHERE is_active=true ORDER BY module, sequence")}

class ApprovalFlowReq(BaseModel):
    module: str
    branch_id: str
    min_amount: float = 0
    max_amount: Optional[float] = None
    approver_role: str
    sequence: int = 1

@router.post("/approval-flow")
def create_approval_flow(req: ApprovalFlowReq):
    return execute_returning(
        "INSERT INTO approval_flow (module, branch_id, min_amount, max_amount, approver_role, sequence) VALUES (%s,%s,%s,%s,%s,%s) RETURNING *",
        (req.module, req.branch_id, req.min_amount, req.max_amount, req.approver_role, req.sequence)
    )

@router.delete("/approval-flow/{id}")
def delete_approval_flow(id: str):
    execute("UPDATE approval_flow SET is_active=false WHERE id=%s", (id,))
    return {"status": "deleted"}

# ── Financial Period ──────────────────────────────────────────────

@router.get("/financial-period")
def list_financial_periods(branch_id: str = ""):
    if branch_id:
        return {"items": fetch_all("SELECT * FROM financial_period WHERE branch_id=%s ORDER BY year DESC, month DESC", (branch_id,))}
    return {"items": fetch_all("SELECT * FROM financial_period ORDER BY year DESC, month DESC")}

class FinancialPeriodReq(BaseModel):
    branch_id: str
    year: int
    month: int
    status: str = "open"

@router.post("/financial-period")
def create_financial_period(req: FinancialPeriodReq):
    return execute_returning(
        "INSERT INTO financial_period (branch_id, year, month, status) VALUES (%s,%s,%s,%s) RETURNING *",
        (req.branch_id, req.year, req.month, req.status)
    )

# ── Account Mapping ───────────────────────────────────────────────

@router.get("/account-mapping")
def list_account_mappings(module: str = ""):
    if module:
        return {"items": fetch_all("SELECT * FROM account_mapping WHERE module=%s", (module,))}
    return {"items": fetch_all("SELECT * FROM account_mapping ORDER BY module, transaction_type")}

class AccountMappingReq(BaseModel):
    module: str
    transaction_type: str
    debit_account: str
    credit_account: str
    description: Optional[str] = None

@router.post("/account-mapping")
def create_account_mapping(req: AccountMappingReq):
    return execute_returning(
        "INSERT INTO account_mapping (module, transaction_type, debit_account, credit_account, description) VALUES (%s,%s,%s,%s,%s) RETURNING *",
        (req.module, req.transaction_type, req.debit_account, req.credit_account, req.description)
    )

@router.delete("/account-mapping/{id}")
def delete_account_mapping(id: str):
    execute("DELETE FROM account_mapping WHERE id=%s", (id,))
    return {"status": "deleted"}

# ── Cost Center ───────────────────────────────────────────────────

@router.get("/cost-center")
def list_cost_centers(branch_id: str = ""):
    if branch_id:
        return {"items": fetch_all("SELECT * FROM cost_center WHERE is_active=true AND branch_id=%s", (branch_id,))}
    return {"items": fetch_all("SELECT * FROM cost_center WHERE is_active=true")}

class CostCenterReq(BaseModel):
    branch_id: str
    name: str
    code: Optional[str] = None

@router.post("/cost-center")
def create_cost_center(req: CostCenterReq):
    return execute_returning("INSERT INTO cost_center (branch_id, name, code) VALUES (%s,%s,%s) RETURNING *", (req.branch_id, req.name, req.code))

@router.delete("/cost-center/{id}")
def delete_cost_center(id: str):
    execute("UPDATE cost_center SET is_active=false WHERE id=%s", (id,))
    return {"status": "deleted"}

# ── Tax Purpose ───────────────────────────────────────────────────

@router.get("/tax-purpose")
def list_tax_purposes():
    return {"items": fetch_all("SELECT * FROM tax_purpose WHERE is_active=true")}

class TaxPurposeReq(BaseModel):
    name: str
    rate: float = 0

@router.post("/tax-purpose")
def create_tax_purpose(req: TaxPurposeReq):
    return execute_returning("INSERT INTO tax_purpose (name, rate) VALUES (%s,%s) RETURNING *", (req.name, req.rate))

@router.delete("/tax-purpose/{id}")
def delete_tax_purpose(id: str):
    execute("UPDATE tax_purpose SET is_active=false WHERE id=%s", (id,))
    return {"status": "deleted"}

# ── Cancel Reason ─────────────────────────────────────────────────

@router.get("/cancel-reason")
def list_cancel_reasons(module: str = ""):
    if module:
        return {"items": fetch_all("SELECT * FROM cancel_reason WHERE is_active=true AND module=%s", (module,))}
    return {"items": fetch_all("SELECT * FROM cancel_reason WHERE is_active=true")}

class CancelReasonReq(BaseModel):
    module: str
    reason: str

@router.post("/cancel-reason")
def create_cancel_reason(req: CancelReasonReq):
    return execute_returning("INSERT INTO cancel_reason (module, reason) VALUES (%s,%s) RETURNING *", (req.module, req.reason))

@router.delete("/cancel-reason/{id}")
def delete_cancel_reason(id: str):
    execute("UPDATE cancel_reason SET is_active=false WHERE id=%s", (id,))
    return {"status": "deleted"}

# ── Product Supplier ──────────────────────────────────────────────

@router.get("/product-supplier")
def list_product_suppliers(product_id: str = ""):
    if product_id:
        return {"items": fetch_all("SELECT * FROM product_supplier WHERE product_id=%s", (product_id,))}
    return {"items": fetch_all("SELECT * FROM product_supplier")}

class ProductSupplierReq(BaseModel):
    product_id: str
    supplier_name: str
    supplier_code: Optional[str] = None
    lead_time_days: int = 0
    min_order_qty: int = 1

@router.post("/product-supplier")
def create_product_supplier(req: ProductSupplierReq):
    return execute_returning(
        "INSERT INTO product_supplier (product_id, supplier_name, supplier_code, lead_time_days, min_order_qty) VALUES (%s,%s,%s,%s,%s) RETURNING *",
        (req.product_id, req.supplier_name, req.supplier_code, req.lead_time_days, req.min_order_qty)
    )

# ── Product Batch ─────────────────────────────────────────────────

@router.get("/product-batch")
def list_product_batches(product_id: str = "", branch_id: str = ""):
    conditions = []
    params = []
    if product_id:
        conditions.append("product_id = %s")
        params.append(product_id)
    if branch_id:
        conditions.append("branch_id = %s")
        params.append(branch_id)
    where = "WHERE " + " AND ".join(conditions) if conditions else ""
    return {"items": fetch_all(f"SELECT pb.*, p.name as product_name FROM product_batch pb JOIN product p ON pb.product_id=p.id {where} ORDER BY received_date DESC", tuple(params))}

class ProductBatchReq(BaseModel):
    product_id: str
    branch_id: str
    batch_no: str
    expiry_date: Optional[str] = None
    qty: float = 0
    cost_per_unit: float = 0

@router.post("/product-batch")
def create_product_batch(req: ProductBatchReq):
    return execute_returning(
        "INSERT INTO product_batch (product_id, branch_id, batch_no, expiry_date, qty, cost_per_unit) VALUES (%s,%s,%s,%s,%s,%s) RETURNING *",
        (req.product_id, req.branch_id, req.batch_no, req.expiry_date, req.qty, req.cost_per_unit)
    )

# ── Currency ──────────────────────────────────────────────────────

@router.get("/currency")
def list_currencies():
    return {"items": fetch_all("SELECT * FROM currency WHERE is_active=true")}

# ── Treatment Package ─────────────────────────────────────────────

@router.get("/treatment-package")
def list_treatment_packages(q: str = "", offset: int = 0, limit: int = 50):
    return list_items("treatment_package", q=q, offset=offset, limit=limit)

class TreatmentPackageReq(BaseModel):
    name: str
    description: Optional[str] = None
    total_sessions: int = 1
    price: float = 0

@router.post("/treatment-package")
def create_treatment_package(req: TreatmentPackageReq):
    return execute_returning(
        "INSERT INTO treatment_package (name, description, total_sessions, price) VALUES (%s,%s,%s,%s) RETURNING *",
        (req.name, req.description, req.total_sessions, req.price)
    )

@router.put("/treatment-package/{id}")
def update_treatment_package(id: str, req: TreatmentPackageReq):
    return execute_returning(
        "UPDATE treatment_package SET name=%s, description=%s, total_sessions=%s, price=%s WHERE id=%s RETURNING *",
        (req.name, req.description, req.total_sessions, req.price, id)
    )

@router.delete("/treatment-package/{id}")
def delete_treatment_package(id: str):
    execute("UPDATE treatment_package SET is_active=false WHERE id=%s", (id,))
    return {"status": "deleted"}
